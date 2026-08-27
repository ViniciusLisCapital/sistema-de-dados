"""
Connector para o BEA (Bureau of Economic Analysis) — tabelas NIPA da Secao 2
(Personal Income and Outlays), no formato de "underlying detail".

--------------------------------------------------------------------------------
DUAS PORTAS PARA O MESMO DADO: XLSX E API
--------------------------------------------------------------------------------
    xlsx  https://apps.bea.gov/national/Release/XLS/Underlying/Section2All_xls.xlsx
    api   https://apps.bea.gov/api/data  (dataset NIUnderlyingDetail)

O xlsx tem 12 MB, 22 abas, e nao pede chave nem tem cota. A API pede chave
(`BEA_API_KEY` no `.env`; registro gratuito em https://apps.bea.gov/API/signup/).

A escolha inicial pelo xlsx foi por CONVENIENCIA, nao por robustez -- so nao pedia
chave. Com a chave instalada em 2026-08-26 as duas portas foram medidas uma contra a
outra, e o resultado divide o problema em dois:

**Nos VALORES a API e melhor, e as duas concordam exatamente.** `tests/test_bea_api.py`
confere valor a valor: **608.442 observacoes** (as duas tabelas, 1959-01 -> hoje),
**zero diferentes, diferenca maxima 0**, nada existindo so de um lado, e os rotulos
identicos depois da mesma `_limpar_rotulo()` (a API tambem traz a referencia cruzada
"(55)" no rotulo -- ela nao vem mais limpa). A API e melhor porque entrega numero
tipado: nao depende de `"Line"` na celula A8, de 2 espacos por nivel, de `.....` como
ausente nem de nota de rodape no fim da coluna A. Tudo isso e camada de apresentacao,
e um reformat cosmetico do BEA quebra este parser -- e por isso que ele levanta em vez
de adivinhar. A conferencia entre as duas fontes e o que transforma esse risco em algo
medido, do mesmo jeito que `connectors/bls.py` faz entre a API do BLS e o arquivo
bruto.

**Na ESTRUTURA a API nao serve: ela nao publica hierarquia nenhuma.** Medido no
registro de `GetData`: 10 campos -- TableName, SeriesCode, LineNumber,
LineDescription, TimePeriod, METRIC_NAME, CL_UNIT, UNIT_MULT, DataValue, NoteRef -- e
nenhum e pai, nivel ou indentacao. `LineDescription` vem SEM os espacos da coluna B,
e `LineNumber` e ordem na tabela, nao profundidade. A hierarquia existe so na camada
de apresentacao, e e ela que faz a aba de PCE existir -- drill-down, contribuicao e o
teste de particao todos dependem de pai. Por isso o desenho e HIBRIDO e nao uma troca
de fonte, e por isso `TabelaNipa.fonte` existe: `inflc_pce_dim` exige `"xlsx"`.

Duas divergencias entre o guia oficial (69 paginas, abr/2026) e a API de verdade,
ambas medidas: o campo e `METRIC_NAME` em maiusculas (o guia escreve `Metric_Name`) e
ha um decimo campo, `NoteRef`, que o guia nao lista. E a palavra "vintage" nao aparece
uma vez no guia -- uma versao anterior desta nota afirmava que a chave serviria para
vintages, e nao havia base para isso.

**Limites.** Documentados: 100 requisicoes/min, 100 MB/min, 30 erros/min, com timeout
de 1 minuto. Medido: `Year=X` traz a serie inteira numa requisicao -- 75 MB, 303.410
registros, 10-20s -- e as duas tabelas seguidas (150 MB em ~40s) passaram sem
estrangulamento, mesmo acima do limite anunciado de MB/min. O xlsx nao tem cota
nenhuma; e a unica dimensao em que ele ganha.

--------------------------------------------------------------------------------
AS ABAS QUE INTERESSAM
--------------------------------------------------------------------------------
O nome da aba e o numero da tabela sem pontos, mais o sufixo de frequencia.
As 7 tabelas do arquivo, todas SA, com o numero de linhas e de meses medidos:

  aba         tabela   conteudo                                linhas   meses
  U20404-M    2.4.4U   indice de preco, 2017=100                  402     810
  U20405-M    2.4.5U   despesa nominal, US$ mi SAAR               402     810
  U20403-M    2.4.3U   PCE real, indices de quantidade            402     810
  U20406-M    2.4.6U   PCE real, dolares encadeados               402     234
  U20304-M    2.3.4U   indice de preco, corte grosso               46     810
  U20305-M    2.3.5U   despesa nominal, corte grosso               46     810
  U20306-M    2.3.6U   PCE real encadeado, corte grosso            46     234

`-A` e `-Q` sao as mesmas tabelas em anual e trimestral.

Duas coisas para nao confundir:

- **As 2.3.xU nao sao "o corte por funcao"** — sao "by Major Type of Product **and**
  by Major Function", 46 linhas: uma tabela grossa que mistura os dois criterios, nao
  um espelho de 402 linhas por funcao. A arvore detalhada por funcao e a **2.5.x**, e
  ela NAO esta neste arquivo (a nota de rodape da 2.4.4U referencia as linhas da 2.5.4
  justamente porque e outra tabela).
- **As 4 tabelas 2.4.xU compartilham as MESMAS 402 linhas**, entao a arvore montada
  aqui serve para todas. As de dolares encadeados (`.6U`) comecam bem depois (234
  meses, ~2007) — encadeado nao e publicado ate 1959 nesse detalhe.

**As duas abas que este projeto usa casam linha a linha**: 2.4.4U e 2.4.5U tem as
mesmas 402 linhas de dado, na mesma ordem, com o mesmo rotulo e a mesma
indentacao — conferido linha a linha em `inflc_pce_dim`. Ou seja indice de preco e
valor nominal se juntam pelo NUMERO DA LINHA, sem casar nome, que foi a fonte de
todo o trabalho sujo do lado do CPI (ver `inflc_cpi_dim`: cinco itens perdidos por
uma virgula de diferenca no rotulo).

--------------------------------------------------------------------------------
FORMA DO ARQUIVO
--------------------------------------------------------------------------------
    linha 1   titulo da tabela
    linha 2   unidade  ("[Index numbers, 2017=100; ... seasonally adjusted]")
    linha 3   periodo coberto ("Monthly data from 1959M01 to 2026M06")
    linha 5   data de publicacao
    linha 8   cabecalho: col A = "Line", col D em diante = "1959M01", "1959M02", ...
    linha 9+  dados: col A = numero da linha, col B = rotulo INDENTADO,
              col C = codigo da serie BEA, col D+ = valores

Convencoes do arquivo, todas exercitadas pelo parser:

- **Indentacao = hierarquia**, 2 espacos por nivel, na coluna B. E a unica fonte
  de parentesco no arquivo (nao ha coluna de pai).
- **A linha 1 e a excecao**: `Personal consumption expenditures` vem indentada com
  6 espacos, enquanto `Goods` e `Services` vem com 0. E cosmetico do stub head do
  BEA. `indentacao` sai crua daqui; quem monta a arvore trata a raiz (ver
  `inflc_pce_dim`).
- **`.....`** = nao disponivel. Vira ausencia de linha em `observacoes`.
- **`ZZZZZZ`** na coluna C = o BEA nao publica serie para aquela linha (sao os dois
  "nets": Net expenditures abroad, Net foreign travel). O nominal existe, o indice
  de preco nao.
- **Codigos repetem**: 13 codigos aparecem em duas linhas cada (a mesma serie entra
  duas vezes na arvore — `Health care` sob Household consumption e sob
  Market-based PCE, por exemplo). Conferido: os valores sao identicos nas duas
  posicoes. Por isso a chave e a LINHA, nao o codigo.
- **Rodape**: as ultimas linhas tem texto de nota na coluna A. Sao descartadas
  porque o filtro de linha de dado exige que a coluna A seja um numero.

--------------------------------------------------------------------------------
SO EXISTE SA
--------------------------------------------------------------------------------
O mensal do BEA e dessazonalizado; nao ha contrapartida NSA destas tabelas. Nao e
lacuna de carga, e da fonte — e coerente com o uso (o PCE e lido dessazonalizado).
"""

from __future__ import annotations

import datetime as _dt
import http.client
import json
import os
import pathlib
import re
import tempfile
import time
import urllib.parse
import urllib.request
from dataclasses import dataclass

import openpyxl
import pandas as pd

_URL_SECAO2_UNDERLYING = (
    "https://apps.bea.gov/national/Release/XLS/Underlying/Section2All_xls.xlsx"
)

# Onde o xlsx baixado fica. Reaproveitado no mesmo dia: `update_us.py` roda o script
# da dim e o das series na mesma passada, e sem isto baixaria 12 MB duas vezes. O
# cache em memoria (`_wb_cache`) ja resolve isso dentro de um processo; o cache em
# disco existe para uma segunda execucao no mesmo dia nao rebaixar.
#
# No temp do sistema, nao no repositorio: nenhum connector deste projeto grava dado
# baixado dentro da arvore de codigo (`pdet_ftp.py` devolve bytes e deixa o destino
# para quem chama), e um caminho relativo como "data/bea" seria resolvido a partir do
# CWD -- ou seja, escreveria em lugares diferentes dependendo de onde o script roda.
_CACHE_DIR = pathlib.Path(tempfile.gettempdir()) / "lis_bea"

# Marcadores de nota de rodape do BEA, e referencias cruzadas para a linha
# correspondente da tabela 2.5.4 (declaradas na nota do proprio arquivo). Nada disso
# e nome de item; sai do rotulo e o original fica em `rotulo_bruto`.
_NOTA = re.compile(r"\\\d+\\")
_XREF = re.compile(
    r"\s*\((?:parts? of\s+)?\d+(?:[,\s]+(?:and\s+)?(?:parts? of\s+)?\d+)*\)\s*$"
)

_USER_AGENT = "Mozilla/5.0 (LIS Capital macro data pipeline)"

_wb_cache: dict[pathlib.Path, object] = {}


@dataclass
class TabelaNipa:
    """Uma aba mensal de tabela NIPA, ja separada em estrutura e observacoes.

    Attributes:
        aba:          nome da aba ("U20404-M").
        titulo:       titulo publicado ("Table 2.4.4U. Price Indexes for ...").
        unidade:      linha de unidade, entre colchetes, como o BEA escreve.
        periodo:      "Monthly data from 1959M01 to 2026M06".
        publicado_em: "Data published July 30, 2026".
        periodos:     rotulos de periodo na ordem das colunas ("1959M01", ...).
        estrutura:    uma linha por linha da tabela — linha, code, rotulo,
                      rotulo_bruto, indentacao.
        observacoes:  formato longo — linha, date (1o dia do mes), value. Linhas sem
                      valor publicado simplesmente nao aparecem.
        fonte:        "xlsx" ou "api". Importa porque a API nao traz hierarquia:
                      `estrutura["indentacao"]` e nula quando `fonte == "api"`, e
                      quem monta arvore tem de recusar essa fonte.
    """

    aba: str
    titulo: str
    unidade: str
    periodo: str
    publicado_em: str
    periodos: list[str]
    estrutura: pd.DataFrame
    observacoes: pd.DataFrame
    fonte: str = "xlsx"

    @property
    def sazonalidade(self) -> str:
        return "SA" if "seasonally adjusted" in self.unidade.lower() else "NSA"


def baixar_secao2_underlying(force: bool = False) -> pathlib.Path:
    """Baixa (ou reaproveita) o xlsx da Secao 2, underlying detail.

    Args:
        force: baixa de novo mesmo que o arquivo de hoje ja exista.

    Returns:
        Caminho do xlsx local.

    Raises:
        RuntimeError: se a resposta for pequena demais para ser o arquivo (tipico de
                      pagina de erro devolvida com status 200).
    """
    _CACHE_DIR.mkdir(parents=True, exist_ok=True)
    dest = _CACHE_DIR / f"Section2All_Underlying_{_dt.date.today():%Y%m%d}.xlsx"
    if dest.exists() and not force:
        return dest

    req = urllib.request.Request(_URL_SECAO2_UNDERLYING, headers={"User-Agent": _USER_AGENT})
    with urllib.request.urlopen(req, timeout=300) as r:
        conteudo = r.read()
    if len(conteudo) < 1_000_000:
        raise RuntimeError(
            f"o xlsx da Secao 2 veio com {len(conteudo):,} bytes, muito menor que os "
            f"~12 MB esperados -- provavelmente uma pagina de erro. "
            f"URL: {_URL_SECAO2_UNDERLYING}"
        )
    dest.write_bytes(conteudo)
    return dest


def caminho_cache_hoje() -> pathlib.Path | None:
    """O xlsx de hoje ja esta em disco? Sem baixar nada.

    Serve para uma conferencia oportunista: quem carrega pela API pode comparar com
    o xlsx quando ele JA foi baixado na mesma passada (o script da arvore baixa), e
    pular a conferencia quando isso custaria 12 MB so para conferir.

    Returns:
        O caminho, ou None se o arquivo de hoje nao estiver em cache.
    """
    dest = _CACHE_DIR / f"Section2All_Underlying_{_dt.date.today():%Y%m%d}.xlsx"
    return dest if dest.exists() else None


def anos_param(ini: int | None = None, fim: int | None = None) -> str:
    """Monta o parametro `Year` da API.

    A API aceita uma lista explicita de anos, e e por isso que ela e mais barata que
    o xlsx numa carga de rotina: o xlsx traz sempre os 810 meses inteiros (12 MB),
    enquanto aqui se pede exatamente a janela que vai ser gravada.

    Args:
        ini: primeiro ano. None (nos dois) devolve "X" = a serie inteira.
        fim: ultimo ano.

    Returns:
        "X" ou "2024,2025,2026".

    Raises:
        ValueError: se so um dos dois for dado, ou se a janela estiver invertida.
    """
    if ini is None and fim is None:
        return "X"
    if ini is None or fim is None:
        raise ValueError("passe os dois anos, ou nenhum (para a serie inteira)")
    if fim < ini:
        raise ValueError(f"janela invertida: {ini}..{fim}")
    return ",".join(str(a) for a in range(int(ini), int(fim) + 1))


def _workbook(caminho: pathlib.Path):
    if caminho not in _wb_cache:
        _wb_cache[caminho] = openpyxl.load_workbook(caminho, read_only=True, data_only=True)
    return _wb_cache[caminho]


def _limpar_rotulo(bruto: str) -> str:
    return _XREF.sub("", _NOTA.sub("", bruto).rstrip()).strip()


def _mes(periodo: str) -> _dt.date:
    """'1959M01' -> date(1959, 1, 1)."""
    ano, mes = periodo.split("M")
    return _dt.date(int(ano), int(mes), 1)


def ler_tabela(aba: str, caminho: pathlib.Path | None = None) -> TabelaNipa:
    """Le uma aba mensal do xlsx da Secao 2.

    Args:
        aba:     nome da aba, ex. "U20404-M" (2.4.4U mensal).
        caminho: xlsx local. Default: baixa/reaproveita o do dia.

    Returns:
        TabelaNipa com `estrutura` e `observacoes` separadas.

    Raises:
        KeyError:     se a aba nao existir no arquivo.
        RuntimeError: se o cabecalho nao estiver na forma esperada.
    """
    caminho = caminho or baixar_secao2_underlying()
    wb = _workbook(caminho)
    if aba not in wb.sheetnames:
        raise KeyError(f"aba {aba!r} nao existe em {caminho.name}. Abas: {wb.sheetnames}")

    linhas = list(wb[aba].iter_rows(values_only=True))
    if len(linhas) < 10:
        raise RuntimeError(f"aba {aba!r} tem so {len(linhas)} linhas -- formato inesperado")

    cab = linhas[7]
    if str(cab[0]).strip() != "Line":
        raise RuntimeError(
            f"aba {aba!r}: esperava 'Line' na coluna A da linha 8, achei {cab[0]!r}. "
            "O BEA mudou o layout do arquivo."
        )
    periodos = [str(c) for c in cab[3:] if c]
    fora = [p for p in periodos if not re.fullmatch(r"\d{4}M\d{2}", p)]
    if fora:
        raise RuntimeError(f"aba {aba!r}: periodos fora do formato YYYYMnn: {fora[:5]}")

    datas = [_mes(p) for p in periodos]
    estrutura: list[dict] = []
    obs_linha: list[int] = []
    obs_data: list[_dt.date] = []
    obs_valor: list[float] = []
    for r in linhas[8:]:
        if r[0] is None or not str(r[0]).strip().isdigit():
            continue  # rodape: nota de texto na coluna A
        n = int(str(r[0]).strip())
        bruto = r[1] or ""
        estrutura.append({
            "linha": n,
            "code": (str(r[2]).strip() if r[2] else None),
            "rotulo": _limpar_rotulo(bruto),
            "rotulo_bruto": bruto.strip(),
            "indentacao": len(bruto) - len(bruto.lstrip(" ")),
        })
        for d, v in zip(datas, r[3:3 + len(periodos)]):
            if isinstance(v, (int, float)):
                obs_linha.append(n)
                obs_data.append(d)
                obs_valor.append(float(v))

    return TabelaNipa(
        aba=aba,
        titulo=str(linhas[0][0] or "").strip(),
        unidade=str(linhas[1][0] or "").strip(),
        periodo=str(linhas[2][0] or "").strip(),
        publicado_em=str(linhas[4][0] or "").strip(),
        periodos=periodos,
        estrutura=pd.DataFrame(estrutura),
        observacoes=pd.DataFrame({"linha": obs_linha, "date": obs_data, "value": obs_valor}),
        fonte="xlsx",
    )


# Abas usadas pelo ramo de PCE deste projeto.
ABA_PCE_INDICE = "U20404-M"
ABA_PCE_NOMINAL = "U20405-M"
# ==============================================================================
# CAMINHO DE API
# ==============================================================================
TABELA_PCE_INDICE = "U20404"
TABELA_PCE_NOMINAL = "U20405"

# aba do xlsx -> nome da tabela na API. As duas portas para o mesmo dado.
ABA_PARA_TABELA = {
    ABA_PCE_INDICE: TABELA_PCE_INDICE,
    ABA_PCE_NOMINAL: TABELA_PCE_NOMINAL,
}

_URL_API = "https://apps.bea.gov/api/data"

# Valores que a API manda no lugar de um numero, no campo DataValue.
_SEM_VALOR = {"", "...", "(NA)", "(D)", "(L)", "(NM)"}


def _chave() -> str:
    """Le BEA_API_KEY do ambiente.

    Raises:
        RuntimeError: se nao estiver definida, dizendo como resolver.
    """
    k = os.environ.get("BEA_API_KEY", "").strip()
    if not k:
        raise RuntimeError(
            "BEA_API_KEY nao esta no ambiente. O caminho de xlsx (`ler_tabela`) nao "
            "precisa de chave -- so o de API (`ler_tabela_api`). Registro gratuito em "
            "https://apps.bea.gov/API/signup/, e a chave vai no `.env` (ver "
            "`.env.example`)."
        )
    return k


def _get_api(**params) -> dict:
    """Uma chamada a API do BEA, com as duas defesas que a medicao exigiu.

    1. **Erro vem com HTTP 200.** Medido: chave invalida, chave vazia, tabela
       inexistente, frequencia invalida e ano fora do range, todos respondem 200 -- e
       o erro aparece em UM DE DOIS lugares, `BEAAPI.Error` (erro do dataset) ou
       `BEAAPI.Results.Error` (erro de chave). Olhar so um deixa o outro passar como
       resposta valida e vazia.
    2. **`IncompleteRead` acontece e nao e limite de tamanho.** Medido: a serie
       inteira desde 1959 vem numa requisicao de 75 MB sem reclamar, e o mesmo pedido
       que falhou uma vez passou nas tres tentativas seguintes. E truncamento de
       conexao, entao repete em vez de desistir ou de fatiar.

    Args:
        **params: parametros da query, sem UserID nem ResultFormat.

    Returns:
        O no `BEAAPI` da resposta.

    Raises:
        RuntimeError: se a API devolver erro em qualquer um dos dois nos.
    """
    q = {"UserID": _chave(), "ResultFormat": "JSON", **params}
    url = _URL_API + "?" + urllib.parse.urlencode(q)
    for tentativa in range(4):
        try:
            with urllib.request.urlopen(url, timeout=600) as r:
                bruto = r.read()
            break
        except http.client.IncompleteRead:
            if tentativa == 3:
                raise
            time.sleep(2 * (tentativa + 1))

    api = json.loads(bruto)["BEAAPI"]
    # CUIDADO: a resposta ECOA a chave de volta em Request.RequestParam (medido).
    # Nunca colocar o corpo cru numa mensagem de erro nem num log.
    erro = api.get("Error") or (api.get("Results") or {}).get("Error")
    if erro:
        det = (erro.get("ErrorDetail") or {}).get("Description", "")
        alvo = ", ".join(f"{k}={v}" for k, v in params.items())
        raise RuntimeError(
            f"BEA API erro {erro.get('APIErrorCode')}: "
            f"{erro.get('APIErrorDescription')} {det}".strip() + f"  ({alvo})"
        )
    return api


def ler_tabela_api(tabela: str, anos: str = "X", freq: str = "M") -> TabelaNipa:
    """Le uma tabela de underlying detail pela API, no lugar do xlsx.

    Mesmo retorno de `ler_tabela`, com UMA diferenca que importa: `estrutura` volta
    com `indentacao` nula, porque a API nao publica hierarquia nenhuma. Medido no
    registro de `GetData`: 10 campos -- TableName, SeriesCode, LineNumber,
    LineDescription, TimePeriod, METRIC_NAME, CL_UNIT, UNIT_MULT, DataValue, NoteRef
    -- e nenhum e pai, nivel ou indentacao. `LineDescription` volta SEM os espacos de
    indentacao que o xlsx tem na coluna B (conferido no repr) e `LineNumber` e ordem
    dentro da tabela, nao profundidade. Por isso `fonte` existe no retorno e
    `inflc_pce_dim` exige `fonte == "xlsx"`: montar a arvore por aqui e impossivel,
    nao apenas pior.

    Duas coisas que o guia oficial escreve diferente do que a API faz, medidas: o
    campo e `METRIC_NAME` em maiusculas (o guia diz `Metric_Name`) e ha um decimo
    campo, `NoteRef`, que o guia nao lista.

    O que a API tem de melhor e o VALOR: numero tipado, sem depender de `"Line"` na
    celula A8, de 2 espacos por nivel, de `.....` como ausente nem de nota de rodape
    no fim da coluna A. As duas fontes foram conferidas valor a valor -- ver
    `conferir_api_xlsx`.

    Args:
        tabela: nome da tabela na API, ex. "U20404" (2.4.4U) -- ver
                `ABA_PARA_TABELA`.
        anos:   "X" para a serie inteira (default) ou "2024,2025,2026".
        freq:   "M", "Q" ou "A".

    Returns:
        TabelaNipa com `fonte="api"` e `estrutura["indentacao"]` nula.

    Raises:
        RuntimeError: erro da API, ou resposta sem registro nenhum.
    """
    api = _get_api(method="GetData", datasetname="NIUnderlyingDetail",
                   TableName=tabela, Frequency=freq, Year=anos)
    res = api["Results"]
    dados = res.get("Data") or []
    if not dados:
        raise RuntimeError(f"BEA API devolveu 0 registros para {tabela}/{freq}/{anos}")

    # A primeira nota traz titulo, unidade entre colchetes e data de revisao, tudo
    # numa string: "Table 2.4.4U. Price Indexes ... [Index numbers, 2017=100; ...]
    # - LastRevised: August 26, 2026". Conferido contra o xlsx: titulo e unidade saem
    # identicos as linhas 1 e 2 do arquivo, e o LastRevised bate com o
    # "Data published" dele.
    nota = next((n["NoteText"] for n in (res.get("Notes") or [])
                 if n.get("NoteRef") == tabela), "")
    m = re.match(r"^(?P<titulo>[^\[]*?)\s*(?P<unidade>\[[^\]]*\])?"
                 r"(?:\s*-\s*LastRevised:\s*(?P<rev>.*))?$", nota)
    titulo = (m.group("titulo") or "").strip() if m else nota.strip()
    unidade = (m.group("unidade") or "").strip() if m else ""
    revisao = (m.group("rev") or "").strip() if m else ""

    est: dict[int, dict] = {}
    obs_linha: list[int] = []
    obs_data: list[_dt.date] = []
    obs_valor: list[float] = []
    periodos: set[str] = set()
    for r in dados:
        n = int(r["LineNumber"])
        periodos.add(r["TimePeriod"])
        if n not in est:
            bruto = r["LineDescription"]
            est[n] = {
                "linha": n,
                "code": (r.get("SeriesCode") or "").strip() or None,
                "rotulo": _limpar_rotulo(bruto),
                "rotulo_bruto": bruto.strip(),
                "indentacao": None,  # a API nao tem hierarquia -- ver docstring
            }
        v = (r.get("DataValue") or "").replace(",", "").strip()
        if v in _SEM_VALOR:
            continue
        obs_linha.append(n)
        obs_data.append(_mes(r["TimePeriod"]))
        obs_valor.append(float(v))

    ordenados = sorted(periodos)
    return TabelaNipa(
        aba=tabela,
        titulo=titulo,
        unidade=unidade,
        periodo=(f"{freq} data from {ordenados[0]} to {ordenados[-1]}"
                 if ordenados else ""),
        publicado_em=(f"Data published {revisao}" if revisao else ""),
        periodos=ordenados,
        estrutura=pd.DataFrame([est[k] for k in sorted(est)]),
        observacoes=pd.DataFrame({"linha": obs_linha, "date": obs_data,
                                  "value": obs_valor}),
        fonte="api",
    )


def indexar_obs(obs: pd.DataFrame, anos: tuple[int, int] | None = None) -> dict:
    """`observacoes` -> {(linha, date): value}, opcionalmente recortado por ano.

    Existe para que as comparacoes entre as duas portas usem UMA indexacao, e nao uma
    por lugar que compara. `date` sai sempre como `datetime.date`, venha a coluna como
    date ou como Timestamp do pandas -- se cada chamador normalizasse por conta, uma
    comparacao entre um lado com date e outro com Timestamp daria "nada em comum" sem
    levantar nada.

    Args:
        obs:  DataFrame com colunas linha, date, value.
        anos: (primeiro, ultimo) inclusive, ou None para tudo.

    Returns:
        dict de (linha, date) -> value.
    """
    fora = {}
    for r in obs.itertuples():
        d = r.date.date() if hasattr(r.date, "date") else r.date
        if anos and not (anos[0] <= d.year <= anos[1]):
            continue
        fora[(int(r.linha), d)] = r.value
    return fora


def comparar_obs(da: dict, dx: dict) -> dict:
    """Compara duas indexacoes de observacoes.

    Args:
        da: um lado (por convencao, a API).
        dx: o outro (por convencao, o xlsx).

    Returns:
        dict com `n_comum`, `n_so_a`, `n_so_b`, `n_diferentes`, `dif_max`, `onde_dif`.
    """
    comuns = da.keys() & dx.keys()
    difs = [(abs(da[k] - dx[k]), k) for k in comuns]
    pior, onde = max(difs) if difs else (0.0, None)
    return {
        "n_comum": len(comuns),
        "n_so_a": len(da.keys() - dx.keys()),
        "n_so_b": len(dx.keys() - da.keys()),
        "n_diferentes": sum(1 for d, _ in difs if d > 0),
        "dif_max": pior,
        "onde_dif": onde,
    }


def conferir_api_xlsx(aba: str, caminho: pathlib.Path | None = None,
                      anos: tuple[int, int] | None = None) -> dict:
    """Confere as duas portas para a mesma tabela, valor a valor.

    E o teste de aceitacao do caminho de API, e existe porque as duas fontes sao
    independentes -- parser de planilha de um lado, JSON tipado do outro. Onde elas
    concordam nao ha erro de leitura em nenhuma das duas. Mesma ideia da conferencia
    que `connectors/bls.py` faz entre a API do BLS e o arquivo bruto.

    Args:
        aba:     aba do xlsx, ex. "U20404-M". A tabela da API sai de
                 `ABA_PARA_TABELA` e a frequencia do sufixo da aba.
        caminho: xlsx local. Default: baixa/reaproveita o do dia.
        anos:    (primeiro, ultimo) para conferir so uma janela -- o pedido a API sai
                 recortado tambem, entao uma janela curta e barata. None = tudo.

    Returns:
        dict com o resultado de `comparar_obs` mais `aba`, `tabela_api`,
        `n_linhas_api`, `n_linhas_xlsx`, `n_rotulos_diferentes`, `n_codigos_diferentes`
        e as datas de publicacao das duas.

    Raises:
        KeyError: se a aba nao tiver tabela correspondente na API.
    """
    if aba not in ABA_PARA_TABELA:
        raise KeyError(f"aba {aba!r} nao tem tabela mapeada na API. "
                       f"Conhecidas: {sorted(ABA_PARA_TABELA)}")
    freq = aba.rsplit("-", 1)[-1]
    a = ler_tabela_api(ABA_PARA_TABELA[aba], anos=anos_param(*(anos or (None, None))),
                       freq=freq)
    x = ler_tabela(aba, caminho=caminho)

    fora = comparar_obs(indexar_obs(a.observacoes, anos),
                        indexar_obs(x.observacoes, anos))

    def _dif(col):
        va = dict(zip(a.estrutura["linha"], a.estrutura[col]))
        vx = dict(zip(x.estrutura["linha"], x.estrutura[col]))
        return sum(1 for l, v in va.items() if l in vx and vx[l] != v), len(va), len(vx)

    ndifr, n_api, n_xls = _dif("rotulo")
    ndifc, _, _ = _dif("code")
    fora.update({
        "aba": aba,
        "tabela_api": a.aba,
        "n_linhas_api": n_api,
        "n_linhas_xlsx": n_xls,
        "n_rotulos_diferentes": ndifr,
        "n_codigos_diferentes": ndifc,
        "publicado_api": a.publicado_em,
        "publicado_xlsx": x.publicado_em,
    })
    return fora
