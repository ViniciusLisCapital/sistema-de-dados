"""
Connector para o **anexo estatistico** do Relatorio de Politica Monetaria (RPM)
do BCB -- a planilha xlsx que publica os dados por tras de cada grafico e tabela
do relatorio.

Diferente de todos os outros connectors do projeto, aqui a unidade nao e uma
serie: e uma EDICAO. Cada trimestre o BCB publica um arquivo novo que republica
a serie inteira revisada, entao o mesmo trimestre do calendario aparece com
valores diferentes em edicoes diferentes. E essa dimensao -- o vintage -- que da
sentido ao connector: nenhuma outra fonte do projeto diz o que o BCB *achava* na
epoca, so o que ele acha hoje.

Fonte: https://www.bcb.gov.br/content/ri/relatorioinflacao/<AAAAMM>/<prefixo><AAAAMM>anp.xlsx

## Prefixo do arquivo muda no meio da serie

O relatorio se chamava "Relatorio de Inflacao" (RI) ate dezembro/2024 e virou
"Relatorio de Politica Monetaria" (RPM) em marco/2025; o nome do arquivo
acompanhou (`ri202412anp.xlsx` -> `rpm202503anp.xlsx`). `_prefixo()` resolve pela
data, mas `url_de()` tenta o outro prefixo se o primeiro der 404 -- se o BCB
mudar de nome de novo, a descoberta continua funcionando sem editar codigo.

## Comeco da serie: 2021-09

Confirmado ao vivo (2026-08) varrendo 2014-03 -> 2026-06 nos dois prefixos:
nada antes de 2021-09 responde 200. O relatorio existe desde 1999, mas so passou
a publicar anexo de dados nessa edicao. Antes disso os numeros existem apenas
como imagem de grafico no PDF -- fora de alcance deste connector.

## Listagem de edicoes: existe, e para o PDF

`edicoes()` traz as 109 edicoes de 1999-06 a 2026-06 numa requisicao (`api/servico/sitebcb/rpm/
ultimas`), com data de publicacao e URL do PDF. Isso resolve a descoberta do RELATORIO, nao a do
anexo: a listagem nao diz nada sobre o xlsx, e nem toda edicao listada tem anexo (so de 2021-09 em
diante). Por isso a enumeracao abaixo continua existindo para o anexo.

## Sem listagem de diretorio (para o anexo)

A pasta nao lista conteudo e a pagina do relatorio e SPA (requests traz so o
shell). Entao `vintages_disponiveis()` ENUMERA os trimestres candidatos desde
2021-09 e testa cada URL com um GET de 2 bytes (`Range: bytes=0-1`). HEAD nao e
usado de proposito: nem todo caminho do CDN do BCB responde a HEAD, e o range
GET foi o que se confirmou estavel.

## As abas mudam de nome entre edicoes

O numero do grafico anda a cada edicao (o hiato do produto ja foi `Graf 2.2.3`,
`2.2.4`, `2.2.6` e `2.2.8`), entao localizar aba por NOME e furado. Use
`localizar_aba()`, que casa um regex contra o bloco de cabecalho da coluna A --
onde vive o titulo publicado ("Grafico 2.2.8 - Hiato do produto: estimativas e
dispersao"). O titulo tambem muda, mas muito menos, e o regex absorve.

Exemplo de uso:

    from connectors.bcb_rpm import AnexoRPM

    anexo = AnexoRPM()
    vintages = anexo.vintages_disponiveis()          # [date(2021,9,1), ...]
    wb = anexo.abrir(vintages[-1])                   # openpyxl, read-only
    ws, titulo = anexo.localizar_aba(wb, r"grafico 2\\.2\\.\\d+ .*hiato do produto")
    grade = anexo.grade(ws)                          # DataFrame cru, header=None
"""

from __future__ import annotations

import datetime as dt
import io
import re
import time
import unicodedata
from dataclasses import dataclass

import openpyxl
import pandas as pd
import requests

_BASE_URL = "https://www.bcb.gov.br/content/ri/relatorioinflacao"

# Listagem de edicoes. A colecao chama-se `rpm` e devolve a serie INTEIRA desde 1999-06, inclusive
# as edicoes publicadas quando o relatorio ainda se chamava RI -- a colecao irma `ri` e subconjunto
# (para em 2024-12), entao nao serve. Descoberto por tentativa: `relatorioinflacao`,
# `relatoriopoliticamonetaria` e variantes dao 400.
_API_EDICOES = "https://www.bcb.gov.br/api/servico/sitebcb/rpm"

# Primeira edicao com anexo estatistico (ver docstring do modulo).
PRIMEIRO_VINTAGE = dt.date(2021, 9, 1)

# O relatorio virou "RPM" nesta edicao; antes era "RI".
_VIRADA_RPM = dt.date(2025, 3, 1)

_RE_WS = re.compile(r"\s+")

# Quantas linhas do topo da coluna A formam o bloco de cabecalho (capitulo,
# secao, titulo, fonte, nota). Medido: o titulo nunca passa da 4a linha.
_LINHAS_CABECALHO = 8


def normalizar(texto: object) -> str:
    """Minusculas, sem acento, espacos colapsados -- para casar titulo de aba."""
    s = unicodedata.normalize("NFD", str(texto).lower().replace("±", "+-"))
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    return _RE_WS.sub(" ", s).strip()


def _prefixo(vintage: dt.date) -> str:
    return "rpm" if vintage >= _VIRADA_RPM else "ri"


def trimestres_desde(inicio: dt.date = PRIMEIRO_VINTAGE,
                     fim: dt.date | None = None) -> list[dt.date]:
    """Meses de publicacao candidatos (mar/jun/set/dez) no intervalo."""
    fim = fim or dt.date.today()
    out, ano = [], inicio.year
    while ano <= fim.year:
        for mes in (3, 6, 9, 12):
            d = dt.date(ano, mes, 1)
            if inicio <= d <= fim:
                out.append(d)
        ano += 1
    return out


@dataclass(frozen=True)
class Edicao:
    """Uma edicao do relatorio, como a API a devolve."""

    vintage: dt.date  # data de publicacao (DataReferencia)
    ano_mes: str      # '200006' -- e o que a URL usa, e o identificador natural da edicao
    url_pdf: str
    titulo: str

    @property
    def nome_arquivo(self) -> str:
        return f"rpm_{self.ano_mes}_{self.vintage:%Y-%m-%d}.pdf"


def edicoes(*, quantidade: int = 500, timeout: float = 60.0) -> list[Edicao]:
    """Todas as edicoes publicadas, em ordem cronologica.

    UMA requisicao devolve o arquivo inteiro -- 109 edicoes de 1999-06 a 2026-06 sem buraco
    (medido 2026-08-25). Ver "Listagem de edicoes" na docstring do modulo.
    """
    resp = requests.get(f"{_API_EDICOES}/ultimas",
                        params={"quantidade": quantidade, "filtro": ""},
                        headers={"User-Agent": "Mozilla/5.0"}, timeout=timeout)
    resp.raise_for_status()
    out = []
    for item in resp.json().get("conteudo") or []:
        url = item["Url"]
        m = re.search(r"/(\d{6})/", url)
        if not m:
            continue
        out.append(Edicao(
            vintage=dt.date.fromisoformat(item["DataReferencia"][:10]),
            ano_mes=m.group(1),
            url_pdf=url if url.startswith("http") else f"https://www.bcb.gov.br{url}",
            titulo=item.get("Titulo", "").strip(),
        ))
    return sorted(out, key=lambda e: e.vintage)


def baixar_pdf(edicao: Edicao, *, tentativas: int = 3, timeout: float = 180.0) -> bytes:
    """Baixa o PDF de uma edicao. Retenta com backoff -- o CDN do BCB da timeout esporadico."""
    ultimo: Exception | None = None
    for k in range(tentativas):
        try:
            resp = requests.get(edicao.url_pdf, headers={"User-Agent": "Mozilla/5.0"},
                                timeout=timeout)
            resp.raise_for_status()
            return resp.content
        except requests.RequestException as err:  # noqa: PERF203
            ultimo = err
            time.sleep(2 * (k + 1))
    raise RuntimeError(f"falhou baixar {edicao.url_pdf}: {ultimo}")


class AnexoRPM:
    """Cliente para os anexos estatisticos do RI/RPM do BCB."""

    def __init__(self, *, timeout: float = 120.0):
        self.timeout = timeout
        self._session = requests.Session()
        self._session.headers.update({"User-Agent": "Mozilla/5.0"})
        self._cache_bytes: dict[dt.date, bytes] = {}
        self._cache_url: dict[dt.date, str] = {}

    # ------------------------------------------------------------------ URLs

    def _url(self, vintage: dt.date, prefixo: str) -> str:
        ym = f"{vintage:%Y%m}"
        return f"{_BASE_URL}/{ym}/{prefixo}{ym}anp.xlsx"

    def url_de(self, vintage: dt.date) -> str | None:
        """URL que responde para esta edicao, ou None se nenhuma responder.

        Tenta o prefixo esperado pela data e, se der 404, o outro -- ver
        "Prefixo do arquivo muda no meio da serie" na docstring do modulo.
        """
        if vintage in self._cache_url:
            return self._cache_url[vintage]
        esperado = _prefixo(vintage)
        for prefixo in (esperado, "ri" if esperado == "rpm" else "rpm"):
            url = self._url(vintage, prefixo)
            try:
                resp = self._session.get(url, timeout=self.timeout,
                                         headers={"Range": "bytes=0-1"})
            except requests.RequestException:
                continue
            if resp.status_code in (200, 206):
                self._cache_url[vintage] = url
                return url
        return None

    def vintages_disponiveis(self, *, desde: dt.date = PRIMEIRO_VINTAGE,
                             ate: dt.date | None = None,
                             ignorar: set[dt.date] | None = None) -> list[dt.date]:
        """Edicoes que existem hoje, em ordem cronologica.

        Uma requisicao de 2 bytes por trimestre candidato (~20 e crescendo uma
        por trimestre). Nao ha listagem de diretorio para consultar no lugar.

        Args:
            desde/ate: janela de trimestres candidatos.
            ignorar: edicoes que o chamador ja tem e nao precisa confirmar que
                existem -- entram no resultado sem gastar requisicao. Encolhe a
                rotina de ~20 requisicoes para 1-2 (so os trimestres novos) sem
                perder a deteccao de buraco: uma edicao antiga que falte no
                chamador continua sendo testada.
        """
        ignorar = ignorar or set()
        return [v for v in trimestres_desde(desde, ate)
                if v in ignorar or self.url_de(v)]

    def vintage_mais_recente(self, *, ate: dt.date | None = None) -> dt.date | None:
        """Edicao publicada mais recente, ou None se nenhuma responder.

        Varre de tras para frente e para na primeira que existe: 1-2
        requisicoes em vez das ~20 de `vintages_disponiveis()`. E o que basta
        para quem so quer a serie corrente.
        """
        for vintage in reversed(trimestres_desde(PRIMEIRO_VINTAGE, ate)):
            if self.url_de(vintage):
                return vintage
        return None

    # -------------------------------------------------------------- download

    def download(self, vintage: dt.date) -> bytes:
        """Baixa o xlsx da edicao (~0,4-1,1 MB). Levanta se ela nao existir."""
        if vintage in self._cache_bytes:
            return self._cache_bytes[vintage]
        url = self.url_de(vintage)
        if url is None:
            raise RuntimeError(
                f"anexo estatistico de {vintage:%Y-%m} nao encontrado em {_BASE_URL} "
                f"(testados os prefixos 'ri' e 'rpm'). A serie comeca em "
                f"{PRIMEIRO_VINTAGE:%Y-%m} e so ha edicao em mar/jun/set/dez."
            )
        resp = self._session.get(url, timeout=self.timeout)
        resp.raise_for_status()
        self._cache_bytes[vintage] = resp.content
        return resp.content

    def abrir(self, vintage: dt.date) -> openpyxl.Workbook:
        """Workbook read-only da edicao (150+ abas; ler so a que interessa)."""
        return openpyxl.load_workbook(io.BytesIO(self.download(vintage)),
                                      read_only=True, data_only=True)

    # ----------------------------------------------------------------- abas

    @staticmethod
    def cabecalho(ws) -> list[str]:
        """Linhas nao vazias do topo da coluna A: capitulo, secao, titulo, fonte, nota."""
        linhas = ws.iter_rows(min_row=1, max_row=_LINHAS_CABECALHO, max_col=1,
                              values_only=True)
        return [str(r[0]).strip() for r in linhas if r[0] is not None and str(r[0]).strip()]

    def localizar_aba(self, wb: openpyxl.Workbook, padrao: str) -> tuple[object, str]:
        """Aba cujo bloco de cabecalho casa `padrao` (regex, ja normalizado).

        Casa contra o TITULO publicado, nao contra o nome da aba -- o numero do
        grafico muda de edicao para edicao (ver docstring do modulo).

        Returns:
            (worksheet, titulo que casou).

        Raises:
            RuntimeError: se nenhuma aba casar, com o padrao usado -- o modo de
            falha esperado quando o BCB renomeia um grafico.
        """
        rx = re.compile(padrao)
        for ws in wb.worksheets:
            for linha in self.cabecalho(ws):
                if rx.search(normalizar(linha)):
                    return ws, linha
        raise RuntimeError(
            f"nenhuma aba com titulo casando /{padrao}/ nesta edicao "
            f"({len(wb.sheetnames)} abas). O BCB pode ter renomeado o grafico -- "
            f"conferir o indice do anexo e ajustar o padrao no script de dominio."
        )

    @staticmethod
    def grade(ws) -> pd.DataFrame:
        """Aba inteira como DataFrame cru (header=None), sem descartar nada.

        O parsing de cabecalho/unidade/bloco de dados fica no script de dominio,
        mesmo padrao de connectors/bcb_tabelas_especiais.py: cada grafico do
        anexo tem uma forma diferente e nao ha ganho em adivinhar aqui.
        """
        return pd.DataFrame([list(r) for r in ws.iter_rows(values_only=True)])
