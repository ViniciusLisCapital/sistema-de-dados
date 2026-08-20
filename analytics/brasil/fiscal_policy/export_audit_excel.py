"""
Auditoria em Excel do Impulso Fiscal (IEG + Impulso via Resultado Primario) -- mostra
as series originais (brutas) lado a lado com cada passo intermediario (acumulado,
%PIB, ajuste sazonal STL) ate o numero final de impulso, com FORMULAS DO EXCEL de
verdade em toda etapa puramente aritmetica (soma movel, razao, diferenca, inversao de
sinal) -- so o proprio ajuste sazonal (STL, uma decomposicao LOESS iterativa,
statsmodels.tsa.seasonal.STL) nao e reproduzivel como formula nativa do Excel; essas
celulas vem coladas como valor (calculadas em Python) e ficam marcadas em amarelo.

Espelha exatamente os calculos de generate_report.py's _ieg_contrib_for_esfera() /
_impulso_quarter_via_stl() / _stl_on_valid_window(), so que sem colapsar para o
resultado final -- cada intermediario fica visivel em sua propria coluna. Construido
para auditar a reescrita de 2026-08 (troca do atalho T/T-sobre-acumulado por STL de
verdade), depois de dois bugs reais terem sido encontrados nessa reescrita (STL sobre
janela com reindex/backfill artificial; div de grafico sem largura/altura no CSS --
ver analytics/brasil/fiscal_policy/CLAUDE.md, Gotchas).

Uso:
    uv run python -c "from analytics.brasil.fiscal_policy.export_audit_excel import run; run()"
    # Saida: reports/fiscal_policy_audit.xlsx
"""
from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from analytics.brasil.fiscal_policy import transforms as tf
from analytics.brasil.fiscal_policy.generate_report import (
    _IEG_MULTIPLICADORES,
    _load_flat,
    _load_table,
    _stl_on_valid_window,
)

_OUT = Path("reports/fiscal_policy_audit.xlsx")

_HEADER_FILL = PatternFill(start_color="1F2853", end_color="1F2853", fill_type="solid")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
_GROUP_FILL = PatternFill(start_color="DDE1E9", end_color="DDE1E9", fill_type="solid")
_GROUP_FONT = Font(bold=True, size=10)
_STL_FILL = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
_TOTAL_FONT = Font(bold=True)
_DATE_FMT = "yyyy-mm-dd"
_NUM_FMT = "#,##0.0000"

_ROW_HEADER = 2
_ROW_DATA0 = 3  # first data row


def _set_group_header(ws, col_start: int, col_end: int, label: str) -> None:
    ws.merge_cells(start_row=1, start_column=col_start, end_row=1, end_column=col_end)
    cell = ws.cell(row=1, column=col_start, value=label)
    cell.font = _GROUP_FONT
    cell.fill = _GROUP_FILL
    cell.alignment = Alignment(horizontal="center")


def _write_dates(ws, dates: list, col: int = 1) -> None:
    ws.cell(row=1, column=col, value="Data").font = _HEADER_FONT
    ws.cell(row=1, column=col).fill = _HEADER_FILL
    ws.cell(row=_ROW_HEADER, column=col, value="")
    for i, d in enumerate(dates):
        cell = ws.cell(row=_ROW_DATA0 + i, column=col, value=d.date() if hasattr(d, "date") else d)
        cell.number_format = _DATE_FMT
    ws.column_dimensions[get_column_letter(col)].width = 12


def _write_values(ws, col: int, header: str, values, fill=None) -> None:
    hcell = ws.cell(row=_ROW_HEADER, column=col, value=header)
    hcell.font = Font(bold=True, size=9)
    hcell.alignment = Alignment(wrap_text=True, vertical="bottom")
    if fill is not None:
        hcell.fill = fill
    for i, v in enumerate(values):
        cell = ws.cell(row=_ROW_DATA0 + i, column=col)
        if v is None or (isinstance(v, float) and pd.isna(v)):
            continue
        cell.value = float(v)
        cell.number_format = _NUM_FMT
        if fill is not None:
            cell.fill = fill
    ws.column_dimensions[get_column_letter(col)].width = 13


def _write_formulas(ws, col: int, header: str, formula_fn, n_rows: int) -> None:
    hcell = ws.cell(row=_ROW_HEADER, column=col, value=header)
    hcell.font = Font(bold=True, size=9)
    hcell.alignment = Alignment(wrap_text=True, vertical="bottom")
    col_letter = get_column_letter(col)
    for i in range(n_rows):
        row = _ROW_DATA0 + i
        formula = formula_fn(row, col_letter)
        if formula is None:
            continue
        cell = ws.cell(row=row, column=col, value=formula)
        cell.number_format = _NUM_FMT
    ws.column_dimensions[col_letter].width = 13


# ── IEG ──────────────────────────────────────────────────────────────────────────
_ESFERAS_IEG = [("geral", "Geral"), ("central", "União"), ("estados", "Estados"), ("municipios", "Municípios")]
_CATEGORIAS_IEG = [
    ("folha", "Folha (×1,32)"),
    ("transferencias", "Transferências (×1,46)"),
    ("investimentos", "Investimentos (×1,66)"),
    ("outras", "Outras (×0,64)"),
]


def _ieg_raw_components(wide: pd.DataFrame, prefix: str) -> dict:
    """Mesma logica de _ieg_contrib_for_esfera() em generate_report.py -- duplicada aqui
    deliberadamente (nao importada) para expor os 4 componentes brutos por nome, em vez
    do resultado ja combinado que aquela funcao retorna."""
    def col(name):
        key = f"{prefix}{name}"
        return wide[key] if key in wide.columns else pd.Series(index=wide.index, dtype=float)

    folha = col("salarios_vencimentos")
    transferencias = col("beneficios_previdenciarios_assistenciais")
    investimentos = col("aquisicao_ativos_nao_financeiros")
    despesa_ajustada = col("gasto") - col("consumo_capital_fixo") - col("juros") - col("transferencias_doacoes")
    outras = despesa_ajustada - folha - transferencias - investimentos
    return {"folha": folha, "transferencias": transferencias, "investimentos": investimentos, "outras": outras}


def _build_ieg_sheets(wb: Workbook, wide: pd.DataFrame, pib_pm: pd.Series, idx: pd.DatetimeIndex) -> None:
    n = len(idx)
    dates = list(idx)

    ws_acum = wb.create_sheet("IEG - Acumulado 4T")
    ws_stl = wb.create_sheet("IEG - Trimestre (STL)")
    for ws in (ws_acum, ws_stl):
        _write_dates(ws, dates)

    col_acum = 2
    col_stl = 2

    # PIB nacional -- denominador compartilhado por todas as esferas/categorias (IEG usa
    # sempre o PIB TOTAL, nao um PIB por esfera -- ver docstring de _ieg_contrib_for_esfera()).
    pib_vals = pib_pm.reindex(idx).tolist()
    _write_values(ws_acum, col_acum, "PIB (R$ mi, trimestral, bruto)", pib_vals)
    pib_col_acum = get_column_letter(col_acum)
    col_acum += 1
    _write_formulas(
        ws_acum, col_acum, "PIB Acum. 4T (TTM, R$ mi)",
        lambda r, c: f"=SUM({pib_col_acum}{r-3}:{pib_col_acum}{r})" if r - 3 >= _ROW_DATA0 else None,
        n,
    )
    pib_ttm_col = get_column_letter(col_acum)
    col_acum += 1

    pib_vals_raw_stl = pib_pm.reindex(idx).tolist()
    _write_values(ws_stl, col_stl, "PIB (R$ mi, trimestral, bruto)", pib_vals_raw_stl)
    pib_col_stl = get_column_letter(col_stl)
    col_stl += 1

    for esfera_key, esfera_label in _ESFERAS_IEG:
        prefix = f"{esfera_key}_"
        componentes = _ieg_raw_components(wide, prefix)

        acum_start_col, stl_start_col = col_acum, col_stl
        acum_contrib_cols, stl_contrib_cols = [], []

        for cat_key, cat_label in _CATEGORIAS_IEG:
            mult = _IEG_MULTIPLICADORES[cat_key]
            raw = componentes[cat_key].reindex(idx)

            # ── Acumulado 4T (TTM/TTM), lag=4 -- linear, reconcilia exatamente ──
            raw_col = get_column_letter(col_acum)
            _write_values(ws_acum, col_acum, f"{cat_label}\nBruto (R$ mi)", raw.tolist())
            col_acum += 1
            ttm_col = get_column_letter(col_acum)
            _write_formulas(
                ws_acum, col_acum, f"{cat_label}\nAcum. 4T (R$ mi)",
                lambda r, c, rc=raw_col: f"=SUM({rc}{r-3}:{rc}{r})" if r - 3 >= _ROW_DATA0 else None,
                n,
            )
            col_acum += 1
            pctpib_col = get_column_letter(col_acum)
            _write_formulas(
                ws_acum, col_acum, f"{cat_label}\n% PIB (Acum. 4T)",
                lambda r, c, tc=ttm_col: f"={tc}{r}/{pib_ttm_col}{r}*100",
                n,
            )
            col_acum += 1
            contrib_col = get_column_letter(col_acum)
            _write_formulas(
                ws_acum, col_acum, f"{cat_label}\nContrib. (p.p. × {mult})",
                lambda r, c, pc=pctpib_col, m=mult: (
                    f"=({pc}{r}-{pc}{r-4})*{m}" if r - 4 >= _ROW_DATA0 else None
                ),
                n,
            )
            acum_contrib_cols.append(contrib_col)
            col_acum += 1

            # ── Trimestre (STL), lag=1 sobre a serie dessazonalizada ──
            raw_col_s = get_column_letter(col_stl)
            _write_values(ws_stl, col_stl, f"{cat_label}\nBruto (R$ mi)", raw.tolist())
            col_stl += 1
            pctpib_raw_col = get_column_letter(col_stl)
            _write_formulas(
                ws_stl, col_stl, f"{cat_label}\n% PIB (Trimestral)",
                lambda r, c, rc=raw_col_s: f"={rc}{r}/{pib_col_stl}{r}*100",
                n,
            )
            col_stl += 1
            # STL fitado na janela NATIVA do componente (wide.index, sem restringir a idx
            # primeiro) -- igual a _ieg_contrib_for_esfera() em generate_report.py. Fazer
            # reindex(idx) ANTES do STL, aqui, reproduziria exatamente o bug ja corrigido
            # naquela funcao (plato artificial de backfill mudando o fit inteiro) -- ver
            # CLAUDE.md Gotchas. Alinha para `idx` só DEPOIS de já ter rodado o STL.
            raw_native = componentes[cat_key]
            pctgdp_native = raw_native / pib_pm.reindex(raw_native.index) * 100
            sa = _stl_on_valid_window(pctgdp_native, period=4).reindex(idx)
            sa_col = get_column_letter(col_stl)
            _write_values(ws_stl, col_stl, f"{cat_label}\n% PIB Dessaz. (STL)", sa.tolist(), fill=_STL_FILL)
            col_stl += 1
            contrib_col_s = get_column_letter(col_stl)
            _write_formulas(
                ws_stl, col_stl, f"{cat_label}\nContrib. (p.p. × {mult})",
                lambda r, c, sc=sa_col, m=mult: (
                    f"=({sc}{r}-{sc}{r-1})*{m}" if r - 1 >= _ROW_DATA0 else None
                ),
                n,
            )
            stl_contrib_cols.append(contrib_col_s)
            col_stl += 1

        # ── Total da esfera = soma das 4 contribuições (exato por construção, em AMBAS variantes) ──
        total_acum_col = col_acum
        _write_formulas(
            ws_acum, total_acum_col, f"{esfera_label}\nIEG (Acum. 4T)",
            lambda r, c, cols=tuple(acum_contrib_cols): "=" + "+".join(f"{cc}{r}" for cc in cols),
            n,
        )
        for i in range(n):
            ws_acum.cell(row=_ROW_DATA0 + i, column=total_acum_col).font = _TOTAL_FONT
        col_acum += 1
        _set_group_header(ws_acum, acum_start_col, total_acum_col, esfera_label)

        total_stl_col = col_stl
        _write_formulas(
            ws_stl, total_stl_col, f"{esfera_label}\nIEG (Trimestre)",
            lambda r, c, cols=tuple(stl_contrib_cols): "=" + "+".join(f"{cc}{r}" for cc in cols),
            n,
        )
        for i in range(n):
            ws_stl.cell(row=_ROW_DATA0 + i, column=total_stl_col).font = _TOTAL_FONT
        col_stl += 1
        _set_group_header(ws_stl, stl_start_col, total_stl_col, esfera_label)

    for ws in (ws_acum, ws_stl):
        ws.freeze_panes = ws.cell(row=_ROW_DATA0, column=2)


# ── PB Impulso (fisc_nfsp) ────────────────────────────────────────────────────────
_ESFERAS_PB = [
    ("total", "Total (Setor Público Consolidado)", "resultado_primario"),
    ("governo_federal", "Governo Federal (sem BC)", "resultado_primario_governo_federal"),
    ("banco_central", "Banco Central", "resultado_primario_banco_central"),
    ("estados", "Estados", "resultado_primario_estados"),
    ("municipios", "Municípios", "resultado_primario_municipios"),
    ("empresas_estatais", "Empresas Estatais", "resultado_primario_empresas_estatais"),
]


def _build_pb_sheets(wb: Workbook, nfsp: dict, pib_mensal: dict) -> None:
    s_total = nfsp["resultado_primario_pct_pib_12m"]
    dates_dt = pd.to_datetime(s_total["dates"])
    n = len(dates_dt)

    # A aba do STL usa um grid PROPRIO, a uniao das datas das 6 series de fluxo mensal
    # (1998/1999 -> hoje), e nao o grid de resultado_primario_pct_pib_12m (2002-11 ->
    # hoje) que a aba do Acumulado usa. Motivo: as formulas de Soma 3m e Delta 3m olham
    # para tras DENTRO da propria aba -- num grid que comeca em 2002-11, as 3 primeiras
    # linhas ficariam em branco, mesmo o relatorio publicando valor ali (ele roda o STL
    # sobre a janela nativa de cada serie, que comeca antes). Com o grid nativo, cada
    # valor publicado pelo relatorio tem, na aba, todas as celulas de que sua formula
    # precisa -- que e justamente o ponto de uma auditoria.
    stl_date_set = set()
    for _k, _label, pct_prefix in _ESFERAS_PB:
        fluxo = f"{pct_prefix}_fluxo_mensal" if _k != "total" else "resultado_primario_fluxo_mensal"
        stl_date_set.update(nfsp.get(fluxo, {"dates": []})["dates"])
    stl_dates = sorted(stl_date_set)
    n_stl = len(stl_dates)

    ws_acum = wb.create_sheet("PB Impulso - Acum 12m")
    ws_stl = wb.create_sheet("PB Impulso - Trimestre (STL)")
    _write_dates(ws_acum, list(dates_dt))
    _write_dates(ws_stl, list(pd.to_datetime(stl_dates)))

    col_acum, col_stl = 2, 2

    gdp_by_date = dict(zip(pib_mensal["dates"], pib_mensal["values"]))
    pib_col_stl = get_column_letter(col_stl)
    _write_values(ws_stl, col_stl, "PIB (R$ mi, mensal, bruto)", [gdp_by_date.get(d) for d in stl_dates])
    col_stl += 1
    sum3m_pib_col = get_column_letter(col_stl)
    _write_formulas(
        ws_stl, col_stl, "PIB Soma 3m (R$ mi)",
        lambda r, c, pc=pib_col_stl: f"=SUM({pc}{r-2}:{pc}{r})" if r - 2 >= _ROW_DATA0 else None,
        n_stl,
    )
    col_stl += 1

    for esfera_key, esfera_label, pct_prefix in _ESFERAS_PB:
        pct_col_name = f"{pct_prefix}_pct_pib_12m" if esfera_key != "total" else "resultado_primario_pct_pib_12m"
        fluxo_col_name = f"{pct_prefix}_fluxo_mensal" if esfera_key != "total" else "resultado_primario_fluxo_mensal"

        # ── Acum. 12m (Y/Y) -- linear, reconcilia exatamente ──
        pct_series = nfsp.get(pct_col_name, {"dates": [], "values": []})
        by_date = dict(zip(pct_series["dates"], pct_series["values"]))
        aligned_pct = [by_date.get(d) for d in s_total["dates"]]
        start_col = col_acum
        raw_col = get_column_letter(col_acum)
        _write_values(ws_acum, col_acum, f"{esfera_label}\n% PIB Acum. 12m (BCB, bruto)", aligned_pct)
        col_acum += 1
        delta_col = get_column_letter(col_acum)
        _write_formulas(
            ws_acum, col_acum, f"{esfera_label}\nΔ 12m (p.p.)",
            lambda r, c, rc=raw_col: f"={rc}{r}-{rc}{r-12}" if r - 12 >= _ROW_DATA0 else None,
            n,
        )
        col_acum += 1
        impulso_col = col_acum
        _write_formulas(
            ws_acum, impulso_col, f"{esfera_label}\nImpulso (Acum. 12m)",
            lambda r, c, dc=delta_col: f"=-{dc}{r}",
            n,
        )
        for i in range(n):
            ws_acum.cell(row=_ROW_DATA0 + i, column=impulso_col).font = _TOTAL_FONT
        col_acum += 1
        _set_group_header(ws_acum, start_col, impulso_col, esfera_label)

        # ── Trimestre (STL) -- fluxo mensal bruto, suavizado (soma 3m) antes do STL ──
        flow_series = nfsp.get(fluxo_col_name, {"dates": [], "values": []})
        by_date_f = dict(zip(flow_series["dates"], flow_series["values"]))
        aligned_flow = [by_date_f.get(d) for d in stl_dates]

        stl_start_col = col_stl
        flow_col = get_column_letter(col_stl)
        _write_values(ws_stl, col_stl, f"{esfera_label}\nFluxo Mensal (R$ mi, bruto)", aligned_flow)
        col_stl += 1
        sum3m_col = get_column_letter(col_stl)
        _write_formulas(
            ws_stl, col_stl, f"{esfera_label}\nFluxo Soma 3m (R$ mi)",
            lambda r, c, fc=flow_col: f"=SUM({fc}{r-2}:{fc}{r})" if r - 2 >= _ROW_DATA0 else None,
            n_stl,
        )
        col_stl += 1
        pctgdp_col = get_column_letter(col_stl)
        _write_formulas(
            ws_stl, col_stl, f"{esfera_label}\n% PIB Soma 3m",
            lambda r, c, sc=sum3m_col: f"={sc}{r}/{sum3m_pib_col}{r}*100",
            n_stl,
        )
        col_stl += 1

        # STL fitado na janela NATIVA do fluxo desta esfera (flow_series["dates"], NAO
        # pre-recortada para o grid de s_total) -- igual a _impulso_quarter_via_stl() em
        # generate_report.py. governo_federal/estados/municipios comecam em 1998/1999,
        # anos antes de resultado_primario_pct_pib_12m (2002-11) -- recortar para o grid
        # do total ANTES do STL reproduziria o mesmo tipo de bug ja corrigido no IEG
        # (janela de fit mais curta/diferente por esfera muda o resultado). So alinha
        # para o grid de exibicao (s_total["dates"]) DEPOIS de já ter rodado o STL.
        native_dates = flow_series["dates"]
        native_flow_3m = tf.rolling_sum(flow_series["values"], window=3)
        native_gdp_3m = tf.rolling_sum([gdp_by_date.get(d) for d in native_dates], window=3)
        native_pctgdp = [None if (f is None or g is None) else f / g * 100 for f, g in zip(native_flow_3m, native_gdp_3m)]
        # Via _stl_on_valid_window() (e nao tf.stl_seasonal_adjust() direto) pelo mesmo
        # motivo do lado do IEG: native_pctgdp comeca com 2 meses None (soma movel de 3
        # ainda incompleta) e stl_seasonal_adjust() interpola/backfilla antes do fit. E
        # tambem o que _impulso_quarter_via_stl() faz -- passar por aqui e o que mantem as
        # duas implementacoes numericamente identicas.
        native_sa_series = _stl_on_valid_window(
            pd.Series(native_pctgdp, index=pd.to_datetime(native_dates)), period=12,
        )
        sa_by_date = {d.strftime("%Y-%m-%d"): v for d, v in native_sa_series.items()}
        sa = [sa_by_date.get(d) for d in stl_dates]
        sa_col = get_column_letter(col_stl)
        _write_values(ws_stl, col_stl, f"{esfera_label}\n% PIB Dessaz. (STL)", sa, fill=_STL_FILL)
        col_stl += 1
        delta_col_s = get_column_letter(col_stl)
        _write_formulas(
            ws_stl, col_stl, f"{esfera_label}\nΔ 3m (p.p.)",
            lambda r, c, sc=sa_col: f"={sc}{r}-{sc}{r-3}" if r - 3 >= _ROW_DATA0 else None,
            n_stl,
        )
        col_stl += 1
        impulso_col_s = col_stl
        _write_formulas(
            ws_stl, impulso_col_s, f"{esfera_label}\nImpulso (Trimestre)",
            lambda r, c, dc=delta_col_s: f"=-{dc}{r}",
            n_stl,
        )
        for i in range(n_stl):
            ws_stl.cell(row=_ROW_DATA0 + i, column=impulso_col_s).font = _TOTAL_FONT
        col_stl += 1
        _set_group_header(ws_stl, stl_start_col, impulso_col_s, esfera_label)

    for ws in (ws_acum, ws_stl):
        ws.freeze_panes = ws.cell(row=_ROW_DATA0, column=2)


# ── Reconciliação ──────────────────────────────────────────────────────────────────
def _build_reconciliation_sheet(wb: Workbook) -> None:
    """Soma-das-partes vs. total, via formula que aponta para as OUTRAS abas -- deixa
    visivel, para toda a serie historica (nao so o ultimo mes/trimestre), exatamente o
    residuo que motivou o bug fix desta rodada (ver CLAUDE.md Gotchas)."""
    ws_ieg_acum = wb["IEG - Acumulado 4T"]
    ws_ieg_stl = wb["IEG - Trimestre (STL)"]
    ws_pb_acum = wb["PB Impulso - Acum 12m"]
    ws_pb_stl = wb["PB Impulso - Trimestre (STL)"]

    ws = wb.create_sheet("Reconciliação")
    ws.cell(row=1, column=1, value="Reconciliação — soma das partes vs. total (ver Notas metodológicas)").font = Font(bold=True, size=12)

    def esfera_total_col(ws_src, esfera_label):
        """Acha a coluna do total de uma esfera pelo cabecalho mesclado na linha 1."""
        for col in range(2, ws_src.max_column + 1):
            v = ws_src.cell(row=1, column=col).value
            if v == esfera_label:
                # total = ultima coluna do grupo mesclado -- procura a proxima celula com valor != None (proximo grupo) ou max_column
                merged_ranges = [r for r in ws_src.merged_cells.ranges if r.min_row == 1 and r.min_col == col]
                if merged_ranges:
                    return merged_ranges[0].max_col
        return None

    def n_rows(ws_src):
        return ws_src.max_row - _ROW_DATA0 + 1

    row = 3
    blocks = [
        ("IEG — Acum. 4T (Geral vs. União+Estados+Municípios; exato por construção linear)",
         ws_ieg_acum, ["Geral", "União", "Estados", "Municípios"]),
        ("IEG — Trimestre/STL (Geral vs. União+Estados+Municípios; NÃO exato — STL não é linear)",
         ws_ieg_stl, ["Geral", "União", "Estados", "Municípios"]),
        ("PB Impulso — Acum. 12m (Total vs. soma das 5 esferas; exato por construção linear)",
         ws_pb_acum, ["Total (Setor Público Consolidado)", "Governo Federal (sem BC)", "Banco Central", "Estados", "Municípios", "Empresas Estatais"]),
        ("PB Impulso — Trimestre/STL (Total vs. soma das 5 esferas; NÃO exato — STL não é linear)",
         ws_pb_stl, ["Total (Setor Público Consolidado)", "Governo Federal (sem BC)", "Banco Central", "Estados", "Municípios", "Empresas Estatais"]),
    ]

    for title, ws_src, labels in blocks:
        ws.cell(row=row, column=1, value=title).font = Font(bold=True, size=10)
        row += 1
        cols = [esfera_total_col(ws_src, lab) for lab in labels]
        total_col, part_cols = cols[0], cols[1:]
        src_name = ws_src.title
        headers = ["Data", "Total", "Soma das Partes", "Diferença", "|Diferença|"]
        for j, h in enumerate(headers):
            c = ws.cell(row=row, column=1 + j, value=h)
            c.font = _HEADER_FONT
            c.fill = _HEADER_FILL
        header_row = row
        row += 1
        nr = n_rows(ws_src)
        for i in range(nr):
            src_row = _ROW_DATA0 + i
            date_cell = ws.cell(row=row, column=1, value=f"='{src_name}'!A{src_row}")
            date_cell.number_format = _DATE_FMT
            total_letter = get_column_letter(total_col)
            ws.cell(row=row, column=2, value=f"='{src_name}'!{total_letter}{src_row}").number_format = _NUM_FMT
            sum_formula = "=" + "+".join(f"'{src_name}'!{get_column_letter(pc)}{src_row}" for pc in part_cols)
            ws.cell(row=row, column=3, value=sum_formula).number_format = _NUM_FMT
            ws.cell(row=row, column=4, value=f"=B{row}-C{row}").number_format = _NUM_FMT
            ws.cell(row=row, column=5, value=f"=ABS(D{row})").number_format = _NUM_FMT
            row += 1
        row += 2  # spacer between blocks

    for col, width in zip("ABCDE", (12, 14, 16, 12, 12)):
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A1"


def _build_readme_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Leia-me", 0)
    lines = [
        ("Auditoria — Impulso Fiscal (IEG + Impulso via Resultado Primário)", 14, True),
        ("", None, False),
        ("Gerado por analytics/brasil/fiscal_policy/export_audit_excel.py — espelha exatamente os", 10, False),
        ("cálculos de generate_report.py, sem colapsar para o resultado final: cada passo", 10, False),
        ("intermediário (acumulado, %PIB, ajuste sazonal) tem sua própria coluna.", 10, False),
        ("", None, False),
        ("Células AMARELAS = saída do STL (statsmodels.tsa.seasonal.STL, robust=True),", 10, True),
        ("calculada em Python e colada como valor — uma decomposição LOESS iterativa não é", 10, False),
        ("reproduzível como fórmula nativa do Excel. Todas as outras células calculadas são", 10, False),
        ("FÓRMULAS DE VERDADE do Excel (visíveis/editáveis na barra de fórmulas) — some,", 10, False),
        ("razão, diferença, inversão de sinal.", 10, False),
        ("", None, False),
        ("Abas:", 11, True),
        ("  • IEG - Acumulado 4T — variação do acumulado em 4 trimestres (Y/Y), a leitura", 10, False),
        ("    'oficial' do paper (Resende & Pires). Reconcilia EXATAMENTE (soma linear).", 10, False),
        ("  • IEG - Trimestre (STL) — variação T/T sobre a série dessazonalizada por STL.", 10, False),
        ("    NÃO reconcilia exatamente entre esferas (União+Estados+Municípios vs. Geral)", 10, False),
        ("    — STL não é uma operação linear. Reconcilia exatamente DENTRO de uma esfera", 10, False),
        ("    (soma das 4 categorias = total daquela esfera), pois o total é definido como", 10, False),
        ("    essa soma, não recalculado à parte.", 10, False),
        ("  • PB Impulso - Acum 12m — mesma lógica do IEG Acumulado, fonte BCB/fisc_nfsp.", 10, False),
        ("  • PB Impulso - Trimestre (STL) — fluxo mensal bruto do BCB, suavizado numa soma", 10, False),
        ("    móvel de 3 meses (o fluxo bruto mês a mês é ruidoso demais para STL direto —", 10, False),
        ("    pagamentos pontuais podem oscilar um único mês em R$100bi+), depois STL", 10, False),
        ("    (period=12) e diferença de 3 meses. Esta aba tem um grid de datas MAIS LONGO", 10, False),
        ("    que a de Acum. 12m (começa no início do fluxo mensal, 1998/1999, não em", 10, False),
        ("    2002-11) — de propósito: as fórmulas Soma 3m/Δ 3m olham para trás dentro da", 10, False),
        ("    própria aba, então o grid precisa começar antes do primeiro valor publicado.", 10, False),
        ("", None, False),
        ("Ajuste sazonal: os fatores do STL são estimados só até o ÚLTIMO ANO CIVIL COMPLETO", 10, True),
        ("e aplicados congelados ao ano corrente (re-rodados quando o ano fecha) — ver o", 10, False),
        ("Apêndice do relatório HTML, item 'Ajuste Sazonal (STL)'.", 10, False),
        ("  • Reconciliação — soma das partes vs. total, ano a ano, para as 4 combinações", 10, False),
        ("    acima — mostra ao vivo o resíduo esperado no toggle Trimestre (STL).", 10, False),
        ("", None, False),
        ("Metodologia completa, histórico dos bugs encontrados nesta reescrita (STL sobre", 10, False),
        ("janela com reindex/backfill artificial; div de gráfico sem CSS de tamanho) e", 10, False),
        ("decisões de escopo: analytics/brasil/fiscal_policy/CLAUDE.md (Gotchas + corpo principal).", 10, False),
    ]
    for i, (text, size, bold) in enumerate(lines, start=1):
        cell = ws.cell(row=i, column=1, value=text)
        cell.font = Font(size=size or 10, bold=bold)
    ws.column_dimensions["A"].width = 90


def run(output: str = str(_OUT)) -> None:
    print("Carregando dados...")

    efgg = _load_table("fisc_efgg")
    wide = efgg.pivot(index="date", columns="name", values="value")
    pib = _load_table("atv_pib_valores_correntes")
    pib_pm = pib[pib["name"] == "pib_pm"].set_index("date")["value"].sort_index()
    pib_4t = pib_pm.rolling(4).sum()
    idx_full = wide.index.intersection(pib_4t.dropna().index).sort_values()

    nfsp = _load_flat("fisc_nfsp")
    pib_mensal = _load_flat("atv_pib_mensal")["pib_mensal"]

    wb = Workbook()
    wb.remove(wb.active)  # tira a aba "Sheet" default -- todas as abas sao criadas explicitamente

    print("  Montando abas IEG...")
    _build_ieg_sheets(wb, wide, pib_pm, idx_full)
    print("  Montando abas PB Impulso...")
    _build_pb_sheets(wb, nfsp, pib_mensal)
    print("  Montando aba de reconciliação...")
    _build_reconciliation_sheet(wb)
    _build_readme_sheet(wb)

    out_path = Path(output)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"Auditoria salva: {out_path.resolve()}")


if __name__ == "__main__":
    run()
