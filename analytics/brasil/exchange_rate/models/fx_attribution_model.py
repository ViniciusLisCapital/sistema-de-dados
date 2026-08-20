"""FX cause-attribution model — pilot, generalized across asset managers.

Turns qualitative FX commentary (asset-manager letters) into a numeric time
series: for each month, how much of the narrative attributes BRL/USD moves to
each of a fixed set of causal regimes (fiscal, monetary/carry, politics,
global USD, commodities, risk sentiment, China/EM growth, trade policy,
capital flows/positioning).

This is a manual-extraction pilot, not an automated pipeline: each manager's
claims.csv is hand-extracted (Claude reading each source document directly)
following the rules in fx_attribution_model.md. There is no live
text-to-claim extraction function here yet — load_claims()/aggregate_monthly()
only cover the claims -> monthly matrix step.

Sign convention: direction is the claim's implied effect ON BRL, not on
whatever the claim's subject is. +1 = strongly BRL-appreciation-supportive,
-1 = strongly BRL-depreciation-driving. A claim that "the dollar is
strengthening globally" is scored NEGATIVE (bad for BRL), not positive.

Wired into the FX report ("reports/brasil/FX Report.html") as its own tab, "FX
Attribution (Manager Letters)" (2026-07-29; that tab lived in the separate
reports/ppp_dashboard.html until the two were fused in 2026-08) — see
build_dashboard_payload() below, and generate_report._load_models(), which
calls it and threads the result
into ppp_equilibrium.render()'s fxattr_payload argument alongside the other
six tabs. The Excel export (export_excel(), below) is unchanged and still
the place for the claim-level detail sheet and the two Trends-sheet charts.

Framework, generalized across managers (2026-07-29): the taxonomy
(CATEGORIES), extraction rules, and aggregation/export logic are all
manager-agnostic -- only the source corpus (documents.csv + claims.csv)
differs. Each manager gets its own subfolder under fx_attribution_data/
(e.g. fx_attribution_data/kinea/), holding that manager's documents.csv,
claims.csv, and the derived monthly.csv/fx_attribution.xlsx. Adding a new
manager means hand-extracting its own documents.csv/claims.csv into a new
fx_attribution_data/<manager>/ folder (same two-file schema, same 9-category
taxonomy) and calling run(manager="<manager>") -- no code changes needed.
See "Adding a new manager" in fx_attribution_model.md.
"""

import csv
from pathlib import Path

DATA_ROOT = Path(__file__).parent / "fx_attribution_data"

# Fixed taxonomy, shared across every manager -- see fx_attribution_model.md.
CATEGORIES = [
    {"slug": "fiscal_br", "label": "Fiscal (Brasil)", "color": "#BB9B1D"},
    {"slug": "monetary_br", "label": "Monetary policy / rate differential (Brasil)", "color": "#1F2853"},
    {"slug": "politics_br", "label": "Politics / idiosyncratic (Brasil)", "color": "#8E44AD"},
    {"slug": "global_usd", "label": "Global USD / DXY", "color": "#418791"},
    {"slug": "commodities", "label": "Commodities / terms of trade / external accounts", "color": "#C0392B"},
    {"slug": "risk_sentiment", "label": "Global risk sentiment (risk-on/risk-off)", "color": "#7F8C8D"},
    {"slug": "china_em", "label": "China / EM growth", "color": "#E67E22"},
    {"slug": "trade_policy", "label": "Trade policy / tariffs", "color": "#2C3E50"},
    {"slug": "capital_flows", "label": "Capital flows / positioning (technical)", "color": "#16A085"},
]
CATEGORY_SLUGS = [c["slug"] for c in CATEGORIES]

# Display label per manager slug -- fallback (title-cased, underscores to
# spaces) covers any future manager added without a code change, per this
# module's "no code changes needed" design (see module docstring).
MANAGER_LABELS = {"kinea": "Kinea", "verde_asset": "Verde Asset"}


def _manager_label(manager):
    return MANAGER_LABELS.get(manager, manager.replace("_", " ").title())


def manager_dir(manager):
    """Resolve fx_attribution_data/<manager>/, failing loudly if that
    manager's corpus hasn't been extracted yet (no silent empty-folder
    creation -- a missing manager is a mistake to surface, not paper over)."""
    d = DATA_ROOT / manager
    if not d.is_dir():
        available = sorted(p.name for p in DATA_ROOT.iterdir() if p.is_dir())
        raise FileNotFoundError(
            f"No corpus found for manager={manager!r} (expected {d}). "
            f"Available managers: {available}"
        )
    return d


def load_claims(manager):
    with open(manager_dir(manager) / "claims.csv", encoding="utf-8") as f:
        rows = list(csv.DictReader(f))
    for r in rows:
        r["direction"] = float(r["direction"])
    return rows


def load_documents(manager):
    with open(manager_dir(manager) / "documents.csv", encoding="utf-8") as f:
        return list(csv.DictReader(f))


def aggregate_monthly(claims, documents):
    """One row per month: n_documents, n_claims, and each category's summed
    direction. A category with no claims that month is 0 (correct: sum of
    nothing), NOT missing -- n_documents/n_claims are what distinguish a
    genuinely silent month from one with no source material at all.
    """
    months = sorted({d["month"] for d in documents})
    by_month = {
        m: {"month": m, "n_documents": 0, "n_claims": 0, **{s: 0.0 for s in CATEGORY_SLUGS}}
        for m in months
    }
    for d in documents:
        by_month[d["month"]]["n_documents"] += 1
    for c in claims:
        row = by_month[c["month"]]
        row["n_claims"] += 1
        row[c["category"]] += c["direction"]

    return [by_month[m] for m in months]


def rolling_mean(values, window=3):
    """Expanding-then-rolling mean, equivalent to
    pandas .rolling(window, min_periods=1).mean() -- month 1 is itself,
    month 2 is the mean of months 1-2, month 3+ is the trailing `window`.
    Smooths the noisy/sparse monthly sums for charting."""
    out = []
    for i in range(len(values)):
        lo = max(0, i - window + 1)
        chunk = values[lo : i + 1]
        out.append(sum(chunk) / len(chunk))
    return out


def write_monthly_csv(monthly, manager):
    out_path = manager_dir(manager) / "monthly.csv"
    fieldnames = ["month", "n_documents", "n_claims"] + CATEGORY_SLUGS
    with open(out_path, "w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames)
        w.writeheader()
        w.writerows(monthly)
    return out_path


def export_excel(claims, monthly, manager):
    """Three-sheet workbook: claim-level detail (source of truth), the
    aggregated monthly matrix, and a Trends sheet (3-month rolling average
    per category + two line charts) -- regenerated fresh every run, so it
    replaces the need to hand-rebuild charts after each re-extraction."""
    from openpyxl import Workbook
    from openpyxl.chart import LineChart, Reference
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    out_path = manager_dir(manager) / "fx_attribution.xlsx"

    wb = Workbook()

    ws1 = wb.active
    ws1.title = "Claims"
    claim_cols = ["date", "month", "source", "source_file", "category", "direction", "quote", "note"]
    ws1.append([c.replace("_", " ").title() for c in claim_cols])
    for cell in ws1[1]:
        cell.font = Font(bold=True)
    for c in claims:
        ws1.append([c[col] for col in claim_cols])
    widths1 = [11, 8, 14, 24, 14, 9, 55, 55]
    for i, w in enumerate(widths1, start=1):
        ws1.column_dimensions[get_column_letter(i)].width = w

    ws2 = wb.create_sheet("Monthly")
    monthly_cols = ["month", "n_documents", "n_claims"] + CATEGORY_SLUGS
    header_labels = ["Month", "N Documents", "N Claims"] + [c["label"] for c in CATEGORIES]
    ws2.append(header_labels)
    for cell in ws2[1]:
        cell.font = Font(bold=True)
    for m in monthly:
        ws2.append([m[col] for col in monthly_cols])
    widths2 = [10, 12, 9] + [max(16, len(c["label"]) // 1) for c in CATEGORIES]
    for i, w in enumerate(widths2, start=1):
        ws2.column_dimensions[get_column_letter(i)].width = min(w, 42)

    ws3 = wb.create_sheet("Trends")
    months = [m["month"] for m in monthly]
    n_rows = len(months)
    n_cat = len(CATEGORIES)

    # Table 1 (cols A:J): directional relevance -- signed rolling average,
    # +1 = BRL-supportive, -1 = BRL-negative. "Which way is this regime pointing."
    trend_header = ["Month"] + [c["label"] for c in CATEGORIES]
    ws3.append(trend_header)
    for cell in ws3[1]:
        cell.font = Font(bold=True)
    rolling_by_cat = {s: rolling_mean([m[s] for m in monthly]) for s in CATEGORY_SLUGS}
    for i, month in enumerate(months):
        ws3.append([month] + [rolling_by_cat[s][i] for s in CATEGORY_SLUGS])
    ws3.column_dimensions["A"].width = 10
    for i in range(n_cat):
        ws3.column_dimensions[get_column_letter(i + 2)].width = 14

    chart_dir = LineChart()
    chart_dir.title = "Relevância direcional — 3-month rolling average (signed)"
    chart_dir.y_axis.title = "Rolling score (+BRL-supportive / -BRL-negative)"
    chart_dir.x_axis.title = "Month"
    chart_dir.width = 30
    chart_dir.height = 14
    data_dir = Reference(ws3, min_col=2, max_col=1 + n_cat, min_row=1, max_row=1 + n_rows)
    cats = Reference(ws3, min_col=1, max_col=1, min_row=2, max_row=1 + n_rows)
    chart_dir.add_data(data_dir, titles_from_data=True)
    chart_dir.set_categories(cats)
    ws3.add_chart(chart_dir, "B" + str(n_rows + 4))

    # Table 2 (cols L onward): Grau de relevância -- magnitude/salience,
    # rolling average of |score|. "How much has this regime mattered,
    # regardless of direction" -- an opposing +0.5/-0.5 month still shows up
    # here instead of netting to ~0 the way the directional table would.
    offset = n_cat + 3  # leave one blank column after table 1 (A:J -> col 11 blank, table 2 starts col 12/L)
    ws3.cell(row=1, column=offset, value="Month")
    for j, c in enumerate(CATEGORIES):
        ws3.cell(row=1, column=offset + 1 + j, value=c["label"])
    for cell in ws3[1][offset - 1 : offset - 1 + 1 + n_cat]:
        cell.font = Font(bold=True)
    abs_rolling_by_cat = {s: rolling_mean([abs(m[s]) for m in monthly]) for s in CATEGORY_SLUGS}
    for i, month in enumerate(months):
        ws3.cell(row=2 + i, column=offset, value=month)
        for j, s in enumerate(CATEGORY_SLUGS):
            ws3.cell(row=2 + i, column=offset + 1 + j, value=abs_rolling_by_cat[s][i])
    ws3.column_dimensions[get_column_letter(offset)].width = 10
    for j in range(n_cat):
        ws3.column_dimensions[get_column_letter(offset + 1 + j)].width = 14

    chart_rel = LineChart()
    chart_rel.title = "Grau de relevância (regimes) — 3-month rolling average (magnitude, |score|)"
    chart_rel.y_axis.title = "Rolling |score|"
    chart_rel.x_axis.title = "Month"
    chart_rel.width = 30
    chart_rel.height = 14
    data_rel = Reference(ws3, min_col=offset + 1, max_col=offset + n_cat, min_row=1, max_row=1 + n_rows)
    cats_rel = Reference(ws3, min_col=offset, max_col=offset, min_row=2, max_row=1 + n_rows)
    chart_rel.add_data(data_rel, titles_from_data=True)
    chart_rel.set_categories(cats_rel)
    ws3.add_chart(chart_rel, get_column_letter(offset + 1) + str(n_rows + 4))

    wb.save(out_path)
    return out_path


def build_manager_payload(manager):
    """One manager's slice of the dashboard payload: months, per-category
    monthly sums, both rolling views (signed direction + |magnitude|), raw
    document/claim counts per month, and the full claim-level detail (sorted
    chronologically) for the tab's table. Mirrors export_excel()'s two
    Trends-sheet tables (directional vs. magnitude) rather than introducing a
    third view."""
    claims = load_claims(manager)
    documents = load_documents(manager)
    monthly = aggregate_monthly(claims, documents)
    months = [m["month"] for m in monthly]
    claim_cols = ["date", "month", "source", "source_file", "category", "direction", "quote", "note"]
    return {
        "label": _manager_label(manager),
        "months": months,
        "n_documents": [m["n_documents"] for m in monthly],
        "n_claims": [m["n_claims"] for m in monthly],
        "monthly": {s: [m[s] for m in monthly] for s in CATEGORY_SLUGS},
        "rolling": {s: rolling_mean([m[s] for m in monthly]) for s in CATEGORY_SLUGS},
        "rolling_abs": {s: rolling_mean([abs(m[s]) for m in monthly]) for s in CATEGORY_SLUGS},
        "claims": [{col: c[col] for col in claim_cols} for c in sorted(claims, key=lambda c: c["date"])],
        "totals": {
            "n_documents": sum(m["n_documents"] for m in monthly),
            "n_claims": len(claims),
        },
    }


def build_dashboard_payload():
    """Payload for the FX report's "FX Attribution (Manager
    Letters)" tab -- one entry per manager subfolder actually present under
    fx_attribution_data/ (hand-extracted documents.csv + claims.csv), so a
    newly onboarded manager appears automatically, no code change needed. The
    9-category taxonomy travels once at the top level since it's shared
    across every manager."""
    managers = sorted(p.name for p in DATA_ROOT.iterdir() if p.is_dir())
    return {
        "categories": CATEGORIES,
        "managers": {m: build_manager_payload(m) for m in managers},
    }


def run(manager="kinea"):
    claims = load_claims(manager)
    documents = load_documents(manager)
    monthly = aggregate_monthly(claims, documents)
    csv_path = write_monthly_csv(monthly, manager)
    xlsx_path = export_excel(claims, monthly, manager)
    print(f"Wrote {csv_path}")
    print(f"Wrote {xlsx_path}")
    return monthly


if __name__ == "__main__":
    run()
