"""Gera a planilha de auditoria da Curva de Phillips (versao "flavored", 12m Y/Y).

Especificacao estimada em 2026-08 (ver docstring de `_estimate`):

    FCPI_t = b_cpi * CPI_{t-1} + (1-b_cpi) * EXP_t
             + g_fx   * FXgap_{t-2}
             + c_comm * ICBr_USD_{t-1}
             + d_gap  * Hiato_t

Todas as variaveis em % acumulado em 12 meses, observadas no ultimo mes de cada
trimestre (mar/jun/set/dez). Sem intercepto: os pesos de inercia e expectativa
somam 1, o que garante neutralidade de longo prazo (um choque permanente de 1 p.p.
em todos os condicionantes move a inflacao em exatamente 1 p.p.).

A planilha e continua -- historico e projecao na MESMA aba `Modelo` -- para que as
defasagens sejam simples referencias de linha e o usuario possa auditar celula a
celula. As linhas de projecao leem os condicionantes da aba `Cenario`.

    uv run python -c "from analytics.brasil.monetary_policy.phillips_excel import run; run()"
"""

from __future__ import annotations

import logging
from decimal import Decimal
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from sklearn.linear_model import Ridge
from sklearn.model_selection import TimeSeriesSplit

from connectors.mysql import MySQLDataRequester

logger = logging.getLogger(__name__)

_OUT = Path(__file__).parent / "data" / "curva_phillips_auditoria.xlsx"
_N_PROJ = 12          # trimestres projetados
_ALPHAS = np.logspace(-3, 3, 61)

# paleta LIS
_AZUL, _DOURADO, _VERDE = "1F2853", "BB9B1D", "418791"
_F_HDR = PatternFill("solid", fgColor=_AZUL)
_F_IN = PatternFill("solid", fgColor="FFF4CC")      # celula editavel
_F_OUT = PatternFill("solid", fgColor="E8F1F2")     # celula calculada
_F_HIST = PatternFill("solid", fgColor="F5F6F8")
_THIN = Side(style="thin", color="D0D4DE")
_BOX = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


# ---------------------------------------------------------------------------
# dados
# ---------------------------------------------------------------------------

def _read(db: str, tbl: str) -> pd.DataFrame:
    r = MySQLDataRequester(db, tbl)
    r.connect()
    df = r.request_data()
    r.close_connection()
    for c in df.columns:
        if df[c].apply(lambda x: isinstance(x, Decimal)).any():
            df[c] = df[c].astype(float)
    return df


def load_panel() -> pd.DataFrame:
    """Painel trimestral (fim de trimestre) com tudo em % acumulado em 12m."""
    ipca = _read("macro_brasil", "inflc_agregados")
    focus = _read("macro_brasil", "expc_focus")
    ptax = _read("macro_brasil", "cmb_ptax")
    meta = _read("macro_brasil", "inflc_meta")
    icu = _read("macro_brasil", "comm_icbr_usd")
    hia = _read("macro_brasil", "pm_hiato_produto")

    def yoy_from_rates(name: str) -> pd.Series:
        s = ipca[ipca["name"] == name]
        s = s.set_index(pd.to_datetime(s["date"]))["value"].sort_index().asfreq("MS")
        return ((1 + s / 100).rolling(12).apply(np.prod, raw=True) - 1) * 100

    def yoy_from_level(df: pd.DataFrame, name: str) -> pd.Series:
        s = df[df["name"] == name]
        s = s.set_index(pd.to_datetime(s["date"]))["value"].sort_index().resample("MS").mean()
        return (s / s.shift(12) - 1) * 100

    f = focus[(focus["indicador"] == "IPCA") & (focus["horizonte"] == "12m")
              & (focus["suavizada"] == "S") & (focus["base_calculo"] == 0)]
    exp12 = f.set_index(pd.to_datetime(f["date"]))["mediana"].sort_index().resample("MS").last()

    px = ptax[ptax["name"] == "ptax_venda"]
    px = px.set_index(pd.to_datetime(px["date"]))["value"].sort_index().resample("MS").mean()
    full = pd.date_range("1998-01-01", px.index.max(), freq="MS")
    # a meta e anual; estende-se para frente porque o valor corrente vale ate ser revisto
    mt = (meta.set_index(pd.to_datetime(meta["date"]))["value"].sort_index()
          .resample("MS").ffill().reindex(full).ffill())

    fxgap = (px / px.shift(12) - 1) * 100 - (mt - 2.0)   # desvio da trajetoria PPC
    cu = yoy_from_level(icu, "icbr_usd") - 2.0           # IC-Br USD, desvio da inflacao externa

    h = hia[hia["variavel"] == "central"].copy()
    # a tabela data o hiato no PRIMEIRO mes do trimestre; o painel usa o ultimo
    h.index = pd.to_datetime(h["date"]) + pd.DateOffset(months=2)
    hiato = h["value"].sort_index()

    d = pd.DataFrame({
        "FCPI": yoy_from_rates("ipca_livres"),
        "CPI": yoy_from_rates("ipca"),
        "ADM": yoy_from_rates("ipca_administrado"),
        "EXP": exp12, "FXGAP": fxgap, "CU": cu,
    }).dropna()
    d = d[d.index.month.isin([3, 6, 9, 12])]
    d["HIATO"] = hiato.reindex(d.index)
    return d.dropna()


def livres_weight(d: pd.DataFrame) -> float:
    """Peso dos precos livres no IPCA, implicito nas proprias series de 12m.

    Resolve CPI = w*FCPI + (1-w)*ADM por minimos quadrados sem intercepto, o que
    e mais honesto do que cravar os ~74% de memoria: sai do dado que a planilha usa.
    """
    y = (d["CPI"] - d["ADM"]).values
    x = (d["FCPI"] - d["ADM"]).values
    return float(np.dot(x, y) / np.dot(x, x))


# ---------------------------------------------------------------------------
# estimacao
# ---------------------------------------------------------------------------

def _design(sub: pd.DataFrame):
    y = (sub["FCPI"] - sub["EXP"]).values
    X = np.column_stack([
        (sub["CPI_l1"] - sub["EXP"]).values,
        sub["FXGAP_l2"].values,
        sub["CU_l1"].values,
        sub["HIATO"].values,
    ])
    return X, y


def _estimate(sub: pd.DataFrame) -> dict:
    """Ridge com restricao de homogeneidade, alpha por validacao temporal.

    A restricao entra pela transformacao (FCPI-EXP) ~ (CPI_l1-EXP): o coeficiente
    estimado e o de inercia, e o de expectativa cai como 1 - inercia. Sem intercepto.
    """
    X, y = _design(sub)
    tscv = TimeSeriesSplit(n_splits=5)
    best, best_mse = None, np.inf
    for a in _ALPHAS:
        errs = [np.mean((y[te] - Ridge(alpha=a, fit_intercept=False).fit(X[tr], y[tr]).predict(X[te])) ** 2)
                for tr, te in tscv.split(X)]
        if np.mean(errs) < best_mse:
            best_mse, best = np.mean(errs), a
    co = Ridge(alpha=best, fit_intercept=False).fit(X, y).coef_
    b_cpi = float(co[0])
    fit = (b_cpi * sub["CPI_l1"] + (1 - b_cpi) * sub["EXP"]
           + co[1] * sub["FXGAP_l2"] + co[2] * sub["CU_l1"] + co[3] * sub["HIATO"])
    res = sub["FCPI"] - fit
    return {
        "b_cpi": b_cpi, "b_exp": 1 - b_cpi,
        "g_fx": float(co[1]), "c_comm": float(co[2]), "d_gap": float(co[3]),
        "alpha": float(best), "n": len(sub),
        "r2": float(1 - np.sum(res ** 2) / np.sum((sub["FCPI"] - sub["FCPI"].mean()) ** 2)),
        "rmse": float(np.sqrt(np.mean(res ** 2))),
        "dw": float(np.sum(np.diff(res) ** 2) / np.sum(res ** 2)),
    }


def build_lags(d: pd.DataFrame) -> pd.DataFrame:
    d = d.copy()
    d["CPI_l1"] = d["CPI"].shift(1)
    d["FXGAP_l2"] = d["FXGAP"].shift(2)
    d["CU_l1"] = d["CU"].shift(1)
    return d.dropna()


def samples(d: pd.DataFrame) -> dict[str, pd.DataFrame]:
    return {
        "ex-COVID": pd.concat([d.loc[:"2019-12-31"], d.loc["2021-01-01":]]),
        "janela BCB 03T4-19T4": d.loc[:"2019-12-31"],
        "pos-2012": d.loc["2012-01-01":],
        "amostra cheia": d,
    }


# ---------------------------------------------------------------------------
# planilha
# ---------------------------------------------------------------------------

_COLS = [
    ("A", "Data", 12), ("B", "Tipo", 10),
    ("C", "FCPI obs", 11), ("D", "CPI obs", 11), ("E", "Adm obs", 11),
    ("F", "EXP", 11), ("G", "FX gap", 11), ("H", "IC-Br USD", 12), ("I", "Hiato", 10),
    ("J", "FCPI modelo", 13), ("K", "CPI modelo", 12),
    ("L", "FCPI efetivo", 13), ("M", "CPI efetivo", 12), ("N", "Residuo", 11),
]


def _style_header(ws, row: int, cols):
    for letter, title, width in cols:
        c = ws[f"{letter}{row}"]
        c.value = title
        c.font = Font(bold=True, color="FFFFFF", size=10)
        c.fill = _F_HDR
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _BOX
        ws.column_dimensions[letter].width = width
    ws.row_dimensions[row].height = 28


def _sheet_parametros(wb, fits: dict, w_livres: float, chosen: str, d: pd.DataFrame):
    ws = wb.create_sheet("Parametros")
    ws["A1"] = "Curva de Phillips — precos livres (12m Y/Y, trimestral)"
    ws["A1"].font = Font(bold=True, size=14, color=_AZUL)
    ws["A2"] = ("FCPI(t) = b_cpi·CPI(t-1) + b_exp·EXP(t) + g_fx·FXgap(t-2) "
                "+ c_comm·ICBrUSD(t-1) + d_gap·Hiato(t)")
    ws["A2"].font = Font(italic=True, size=10)
    ws["A3"] = "Sem intercepto. b_cpi + b_exp = 1 por construcao (neutralidade de longo prazo)."
    ws["A3"].font = Font(size=9, color="666666")

    ws["A5"] = "PARAMETROS EM USO"
    ws["A5"].font = Font(bold=True, size=11, color=_DOURADO)
    f = fits[chosen]
    rows = [
        ("b_cpi", f["b_cpi"], "Inercia — peso do IPCA cheio defasado 1 trimestre"),
        ("b_exp", f["b_exp"], "Expectativa — Focus IPCA 12m (= 1 - b_cpi)"),
        ("g_fx", f["g_fx"], "Cambio — desvio da variacao 12m do USD/BRL vs. trajetoria PPC, defasado 2T"),
        ("c_comm", f["c_comm"], "Commodities — IC-Br em USD, variacao 12m menos 2%, defasado 1T"),
        ("d_gap", f["d_gap"], "Hiato do produto do BCB (pm_hiato_produto), contemporaneo"),
        ("w_livres", w_livres, "Peso dos precos livres no IPCA — usado para recompor o cheio"),
    ]
    for i, (name, val, desc) in enumerate(rows, start=6):
        ws[f"A{i}"] = name
        ws[f"A{i}"].font = Font(bold=True, name="Consolas", size=10)
        ws[f"B{i}"] = round(val, 6)
        ws[f"B{i}"].fill = _F_IN
        ws[f"B{i}"].border = _BOX
        ws[f"B{i}"].number_format = "0.0000"
        ws[f"C{i}"] = desc
        ws[f"C{i}"].font = Font(size=9, color="444444")
        wb.defined_names.add(DefinedName(name, attr_text=f"Parametros!$B${i}"))

    ws["A13"] = f"Amostra escolhida: {chosen}"
    ws["A13"].font = Font(bold=True, size=10, color=_VERDE)
    ws["A14"] = ("Para trocar de amostra, copie os valores da tabela abaixo para as celulas amarelas. "
                 "Todo o resto da planilha recalcula sozinho.")
    ws["A14"].font = Font(size=9, color="666666")

    ws["A16"] = "ESTIMATIVAS ALTERNATIVAS"
    ws["A16"].font = Font(bold=True, size=11, color=_DOURADO)
    hdr = ["Amostra", "b_cpi", "b_exp", "g_fx", "c_comm", "d_gap", "n", "R2", "RMSE", "DW", "alpha"]
    for j, t in enumerate(hdr):
        c = ws.cell(row=17, column=1 + j, value=t)
        c.font = Font(bold=True, color="FFFFFF", size=10)
        c.fill = _F_HDR
        c.border = _BOX
    for i, (name, ft) in enumerate(fits.items(), start=18):
        vals = [name, ft["b_cpi"], ft["b_exp"], ft["g_fx"], ft["c_comm"], ft["d_gap"],
                ft["n"], ft["r2"], ft["rmse"], ft["dw"], ft["alpha"]]
        for j, v in enumerate(vals):
            c = ws.cell(row=i, column=1 + j, value=round(v, 6) if isinstance(v, float) else v)
            c.border = _BOX
            if isinstance(v, float):
                c.number_format = "0.0000"
            if name == chosen:
                c.font = Font(bold=True)
    for letter, width in zip("ABCDEFGHIJK", [24, 11, 11, 11, 11, 11, 8, 9, 9, 8, 10]):
        ws.column_dimensions[letter].width = width

    ws["A24"] = "DEFINICAO DAS VARIAVEIS"
    ws["A24"].font = Font(bold=True, size=11, color=_DOURADO)
    defs = [
        ("FCPI", "IPCA precos livres, % acumulado em 12 meses", "macro_brasil.inflc_agregados / ipca_livres"),
        ("CPI", "IPCA cheio, % acumulado em 12 meses", "macro_brasil.inflc_agregados / ipca"),
        ("Adm", "IPCA administrados, % acumulado em 12 meses", "macro_brasil.inflc_agregados / ipca_administrado"),
        ("EXP", "Focus IPCA 12m a frente, mediana suavizada, base 0", "macro_brasil.expc_focus"),
        ("FX gap", "Variacao 12m do USD/BRL menos (meta - 2%)", "macro_brasil.cmb_ptax + inflc_meta"),
        ("IC-Br USD", "Variacao 12m do IC-Br em dolar menos 2%", "macro_brasil.comm_icbr_usd"),
        ("Hiato", "Hiato do produto do BCB, % do produto potencial", "macro_brasil.pm_hiato_produto (central)"),
    ]
    for j, t in enumerate(["Variavel", "Definicao", "Fonte"]):
        c = ws.cell(row=25, column=1 + j, value=t)
        c.font = Font(bold=True, color="FFFFFF", size=10)
        c.fill = _F_HDR
        c.border = _BOX
    for i, (a, b, c_) in enumerate(defs, start=26):
        for j, v in enumerate([a, b, c_]):
            cell = ws.cell(row=i, column=1 + j, value=v)
            cell.border = _BOX
            cell.font = Font(size=9)
    ws.column_dimensions["B"].width = 52
    ws.column_dimensions["C"].width = 46

    ws["A35"] = "AVISOS"
    ws["A35"].font = Font(bold=True, size=11, color="C2381F")
    avisos = [
        "1. pm_hiato_produto tem UMA safra (a edicao mais recente do RPM). O historico inteiro "
        "carrega informacao de hoje, entao serve para ajuste dentro da amostra, nao para backtest.",
        "2. Autocorrelacao residual segue alta (DW ~0.8): as observacoes de 12m se sobrepoem em "
        "9 dos 12 meses. As estatisticas de ajuste superestimam a precisao.",
        "3. O Focus subestimou a inflacao realizada em ~0,66 p.p. em media na amostra. Sem intercepto, "
        "esse vies aparece como residuo medio positivo em vez de ser absorvido.",
        "4. Precos administrados NAO sao modelados: entram como condicionante na aba Cenario e so "
        "servem para recompor o IPCA cheio a partir dos livres.",
    ]
    for i, t in enumerate(avisos, start=36):
        ws[f"A{i}"] = t
        ws[f"A{i}"].font = Font(size=9, color="444444")
        ws[f"A{i}"].alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(f"A{i}:K{i}")
        ws.row_dimensions[i].height = 26
    return ws


def _sheet_cenario(wb, d: pd.DataFrame, proj_dates: list[pd.Timestamp]):
    ws = wb.create_sheet("Cenario")
    ws["A1"] = "Condicionantes da projecao"
    ws["A1"].font = Font(bold=True, size=14, color=_AZUL)
    ws["A2"] = ("Edite as celulas AMARELAS. Cada bloco tem a trajetoria trimestre a trimestre e um "
                "CHOQUE que soma a todos os trimestres de uma vez.")
    ws["A2"].font = Font(italic=True, size=10, color="666666")

    last = d.iloc[-1]
    blocks = [
        ("EXP", "Focus IPCA 12m (%)", float(last["EXP"])),
        ("FXGAP", "FX gap: var. 12m USD/BRL - trajetoria PPC (p.p.)", 0.0),
        ("CU", "IC-Br USD: var. 12m - 2% (p.p.)", 0.0),
        ("HIATO", "Hiato do produto (% do potencial)", float(last["HIATO"])),
        ("ADM", "IPCA administrados 12m (%)", float(last["ADM"])),
    ]

    ws["A4"] = "CHOQUES (somam a toda a trajetoria)"
    ws["A4"].font = Font(bold=True, size=11, color=_DOURADO)
    for j, (key, label, _) in enumerate(blocks):
        col = get_column_letter(2 + j * 2)
        c = ws[f"{col}5"]
        c.value = label
        c.font = Font(bold=True, size=9, color="FFFFFF")
        c.fill = _F_HDR
        c.alignment = Alignment(wrap_text=True, horizontal="center")
        c.border = _BOX
        s = ws[f"{col}6"]
        s.value = 0.0
        s.fill = _F_IN
        s.border = _BOX
        s.number_format = "0.00"
        s.font = Font(bold=True)
        wb.defined_names.add(DefinedName(f"choque_{key.lower()}", attr_text=f"Cenario!${col}$6"))
        ws.column_dimensions[col].width = 15
        ws.column_dimensions[get_column_letter(3 + j * 2)].width = 15
    ws["A5"] = "choque:"
    ws["A5"].font = Font(bold=True, size=9)
    ws["A6"] = "→ p.p."
    ws["A6"].font = Font(size=9, color="666666")
    ws.row_dimensions[5].height = 30

    hdr = 8
    ws[f"A{hdr}"] = "Trimestre"
    ws[f"A{hdr}"].font = Font(bold=True, color="FFFFFF", size=10)
    ws[f"A{hdr}"].fill = _F_HDR
    ws[f"A{hdr}"].border = _BOX
    ws.column_dimensions["A"].width = 13
    for j, (key, _, _) in enumerate(blocks):
        b, e = get_column_letter(2 + j * 2), get_column_letter(3 + j * 2)
        for letter, title in ((b, f"{key} base"), (e, f"{key} efetivo")):
            c = ws[f"{letter}{hdr}"]
            c.value = title
            c.font = Font(bold=True, color="FFFFFF", size=9)
            c.fill = _F_HDR
            c.alignment = Alignment(horizontal="center")
            c.border = _BOX

    for i, dt in enumerate(proj_dates):
        r = hdr + 1 + i
        ws[f"A{r}"] = dt.strftime("%Y-%m")
        ws[f"A{r}"].border = _BOX
        ws[f"A{r}"].font = Font(size=10)
        for j, (key, _, base) in enumerate(blocks):
            b, e = get_column_letter(2 + j * 2), get_column_letter(3 + j * 2)
            cb = ws[f"{b}{r}"]
            cb.value = round(base, 4)
            cb.fill = _F_IN
            cb.border = _BOX
            cb.number_format = "0.00"
            ce = ws[f"{e}{r}"]
            ce.value = f"={b}{r}+choque_{key.lower()}"
            ce.fill = _F_OUT
            ce.border = _BOX
            ce.number_format = "0.00"

    note = hdr + len(proj_dates) + 2
    ws[f"A{note}"] = ("Trajetoria base = ultimo valor observado mantido constante, exceto FX gap e IC-Br USD, "
                      "que partem de zero (cambio e commodities avancando exatamente na trajetoria de "
                      "equilibrio — o cenario neutro, nao 'sem variacao').")
    ws[f"A{note}"].font = Font(size=9, color="444444")
    ws[f"A{note}"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(f"A{note}:K{note}")
    ws.row_dimensions[note].height = 30
    return hdr


def _sheet_modelo(wb, d: pd.DataFrame, proj_dates: list[pd.Timestamp], cen_hdr: int):
    ws = wb.create_sheet("Modelo", 0)
    ws["A1"] = "Curva de Phillips — historico e projecao"
    ws["A1"].font = Font(bold=True, size=14, color=_AZUL)
    ws["A2"] = ("Historico e projecao na mesma tabela: as defasagens sao referencias de linha. "
                "Linhas cinzas = historico observado. Linhas em destaque = projecao (condicionantes vem da aba Cenario).")
    ws["A2"].font = Font(italic=True, size=10, color="666666")

    hrow = 4
    _style_header(ws, hrow, _COLS)
    first = hrow + 1

    for i, (dt, row) in enumerate(d.iterrows()):
        r = first + i
        ws[f"A{r}"] = dt.strftime("%Y-%m")
        ws[f"B{r}"] = "Hist"
        for letter, val in (("C", row["FCPI"]), ("D", row["CPI"]), ("E", row["ADM"]),
                            ("F", row["EXP"]), ("G", row["FXGAP"]), ("H", row["CU"]),
                            ("I", row["HIATO"])):
            c = ws[f"{letter}{r}"]
            c.value = round(float(val), 4)
            c.number_format = "0.00"
        for letter in "ABCDEFGHI":
            ws[f"{letter}{r}"].fill = _F_HIST

    n_hist = len(d)
    proj_first = first + n_hist
    for i, dt in enumerate(proj_dates):
        r = proj_first + i
        cr = cen_hdr + 1 + i
        ws[f"A{r}"] = dt.strftime("%Y-%m")
        ws[f"B{r}"] = "Proj"
        ws[f"B{r}"].font = Font(bold=True, color=_DOURADO)
        ws[f"E{r}"] = f"=Cenario!K{cr}"     # ADM efetivo
        ws[f"F{r}"] = f"=Cenario!C{cr}"     # EXP efetivo
        ws[f"G{r}"] = f"=Cenario!E{cr}"     # FX gap efetivo
        ws[f"H{r}"] = f"=Cenario!G{cr}"     # IC-Br USD efetivo
        ws[f"I{r}"] = f"=Cenario!I{cr}"     # hiato efetivo
        for letter in "EFGHI":
            ws[f"{letter}{r}"].fill = _F_OUT
            ws[f"{letter}{r}"].number_format = "0.00"

    last = first + n_hist + len(proj_dates) - 1
    for r in range(first, last + 1):
        # J: a equacao. M(r-1) = IPCA cheio efetivo do trimestre anterior,
        # G(r-2) = FX gap ha dois trimestres, H(r-1) = commodities ha um trimestre.
        ws[f"J{r}"] = (f"=b_cpi*M{r - 1}+b_exp*F{r}+g_fx*G{r - 2}+c_comm*H{r - 1}+d_gap*I{r}"
                       if r >= first + 2 else "")
        ws[f"K{r}"] = f"=w_livres*J{r}+(1-w_livres)*E{r}" if r >= first + 2 else ""
        ws[f"L{r}"] = f'=IF(B{r}="Hist",C{r},J{r})'
        ws[f"M{r}"] = f'=IF(B{r}="Hist",D{r},K{r})'
        ws[f"N{r}"] = f'=IF(AND(B{r}="Hist",J{r}<>""),C{r}-J{r},"")'
        for letter in "JKLMN":
            c = ws[f"{letter}{r}"]
            c.number_format = "0.00"
            if r >= proj_first:
                c.fill = _F_OUT
                c.font = Font(bold=(letter in "JK"))
        for letter in "ABCDEFGHIJKLMN":
            ws[f"{letter}{r}"].border = _BOX

    ws.freeze_panes = f"C{first}"

    stat = last + 2
    ws[f"A{stat}"] = "AJUSTE NO HISTORICO"
    ws[f"A{stat}"].font = Font(bold=True, size=11, color=_DOURADO)
    hist_last = first + n_hist - 1
    rng = f"N{first + 2}:N{hist_last}"
    obs = f"C{first + 2}:C{hist_last}"
    for i, (lab, formula) in enumerate([
        ("Residuo medio (p.p.)", f"=AVERAGE({rng})"),
        ("RMSE (p.p.)", f"=SQRT(SUMSQ({rng})/COUNT({rng}))"),
        ("R2", f"=1-SUMSQ({rng})/SUMPRODUCT(({obs}-AVERAGE({obs}))^2)"),
    ], start=stat + 1):
        ws[f"A{i}"] = lab
        ws[f"A{i}"].font = Font(size=10)
        c = ws[f"B{i}"]
        c.value = formula
        c.number_format = "0.0000"
        c.fill = _F_OUT
        c.border = _BOX
        c.font = Font(bold=True)
    ws[f"A{stat + 5}"] = ("As duas primeiras linhas do historico ficam vazias porque a equacao precisa "
                          "de FX gap defasado em dois trimestres.")
    ws[f"A{stat + 5}"].font = Font(size=9, color="666666")
    return ws


def run(output: Path | str = _OUT, chosen: str = "ex-COVID", n_proj: int = _N_PROJ) -> Path:
    d_raw = load_panel()
    w = livres_weight(d_raw)
    d = build_lags(d_raw)
    fits = {name: _estimate(sub) for name, sub in samples(d).items()}

    logger.info("peso dos livres no IPCA: %.4f", w)
    for name, f in fits.items():
        logger.info("%-22s b_cpi %.4f b_exp %.4f g_fx %.4f c_comm %.4f d_gap %.4f | R2 %.4f DW %.2f",
                    name, f["b_cpi"], f["b_exp"], f["g_fx"], f["c_comm"], f["d_gap"], f["r2"], f["dw"])

    last = d.index[-1]
    proj_dates = [last + pd.DateOffset(months=3 * (i + 1)) for i in range(n_proj)]

    wb = Workbook()
    wb.remove(wb.active)
    cen_hdr = _sheet_cenario(wb, d, proj_dates)
    _sheet_modelo(wb, d, proj_dates, cen_hdr)
    _sheet_parametros(wb, fits, w, chosen, d)
    wb._sheets = [wb["Modelo"], wb["Cenario"], wb["Parametros"]]

    out = Path(output)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    logger.info("planilha salva em %s", out)
    return out


if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO, format="%(message)s")
    run()
