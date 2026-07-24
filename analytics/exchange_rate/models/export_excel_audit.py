"""
Excel audit workbook for the PPP equilibrium + Bayesian deviation model —
lets the user trace the entire data-treatment pipeline by hand, cell by
cell, using live formulas rather than pasted numbers.

Deliberately does NOT re-estimate the regression in Excel (LINEST/OLS was
considered and rejected 2026-07-23: "it's not necessary to construct the
OLS, just plug the parameters"). Instead, the already-fitted Bayesian
posterior means (from bayesian_results/primary_breakeven_summary.csv, saved
by bayesian_deviation_model.run()) are hardcoded as fixed inputs on the
"Model Parameters" sheet. Everything else — equilibrium, deviation,
differencing, 1-month lag, standardization, contribution math, the
cumulative/level bridge — is mechanical and rebuilt as real formulas so it
can be checked independently of the Python code.

Sheet flow (all data sheets share one date backbone: row 2 = 1994-07,
row 384 = 2026-06, so "row r" means the same month on every sheet):
  README                 — how to read this workbook
  Raw Data               — the original series, pulled as-is (static)
  PPP Equilibrium        — equilibrium(t) and deviation(t), base-month
                           selector (formulas)
  Model Parameters       — Bayesian posterior means (hardcoded, labeled)
                           + standardization stats (formulas, over the
                           same sample the model was fit on)
  Deltas                 — first differences, 1-month lag, z-scores
                           (formulas)
  Decomposition & Bridge — contributions, cumulative sums, and the
                           level (R$) bridge from equilibrium to the
                           actual PTAX rate (formulas)

Usage:
    uv run python -c "from analytics.exchange_rate.models.export_excel_audit import run; run()"
"""

from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from analytics.exchange_rate.models.ppp_equilibrium import _DEFAULT_BASE_MONTH, load_data

_OUTPUT = Path(__file__).parent.parent.parent.parent / "reports" / "ppp_model_audit.xlsx"

# Posterior means, primary_breakeven spec — from
# analytics/exchange_rate/models/bayesian_results/primary_breakeven_summary.csv
# (saved by bayesian_deviation_model.run(); NOT re-estimated in this workbook)
_POSTERIOR = {
    "alpha": 0.219,
    "beta_carry": 0.131,
    "beta_tot": -0.268,
    "beta_breakeven": -0.387,
    "beta_fiscal": 0.520,
}

_NAVY = "1F2853"
_GOLD = "BB9B1D"
_TEAL = "02739B"
_GREY = "7A88A8"
_LINE = "D8DCE6"

_HEADER_FILL = PatternFill("solid", fgColor=_NAVY)
_HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
_PARAM_FILL = PatternFill("solid", fgColor="FBF3D9")
_THIN = Side(style="thin", color=_LINE)
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _style_header(ws, row, n_cols):
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = 32


def _autosize(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def build_workbook() -> Workbook:
    df = load_data()  # index: month-start dates, 1994-07 .. latest; row 2 = df.index[0]
    n = len(df)
    first_row, last_row = 2, 1 + n  # e.g. 2 .. 384

    # Fixed sample-window boundaries for the primary_breakeven spec (data-derived,
    # confirmed 2026-07-23: fiscal-risk (CDS) availability from 2007-12 is the
    # binding constraint; +1 for diff, +1 for lag -> sample starts 2008-02).
    months = list(df.index)
    primary_start_date = pd.Timestamp("2008-02-01")
    primary_end_date = pd.Timestamp("2026-05-01")
    anchor_date = pd.Timestamp("2008-01-01")
    primary_start_row = months.index(primary_start_date) + first_row
    primary_end_row = months.index(primary_end_date) + first_row
    anchor_row = months.index(anchor_date) + first_row

    wb = Workbook()
    wb.remove(wb.active)

    # =====================================================================
    # README
    # =====================================================================
    ws = wb.create_sheet("README")
    ws.sheet_view.showGridLines = False
    lines = [
        ("PPP Equilibrium + Bayesian Deviation Model — Audit Workbook", "title"),
        ("", ""),
        ("Every data sheet in this workbook shares one date backbone: row 2 = "
         f"{months[0].strftime('%Y-%m')}, row {last_row} = {months[-1].strftime('%Y-%m')}. "
         "Row r means the same month on every sheet, so a formula on 'Deltas' row 200 "
         "and a formula on 'Raw Data' row 200 refer to the same month.", "body"),
        ("", ""),
        ("What is NOT re-derived here", "h2"),
        ("The regression coefficients (alpha, beta_carry, beta_tot, beta_breakeven, "
         "beta_fiscal) are the posterior MEANS from the already-fitted Bayesian model "
         "(PyMC/NUTS, primary_breakeven spec), saved to "
         "analytics/exchange_rate/models/bayesian_results/primary_breakeven_summary.csv. "
         "They are hardcoded on the 'Model Parameters' sheet, not re-estimated with "
         "Excel formulas — fitting a Bayesian posterior isn't something a spreadsheet "
         "formula can reproduce, and re-deriving it wasn't the point of this workbook.", "body"),
        ("Everything else IS rebuilt as live formulas: the PPP equilibrium construction, "
         "the log-deviation, the first differences, the 1-month lag, the standardization "
         "(mean/stdev), and the full contribution -> cumulative -> level-bridge math. "
         "All of that is purely mechanical, so it's fully checkable independently of the "
         "Python code.", "body"),
        ("", ""),
        ("Sheet-by-sheet", "h2"),
        ("Raw Data — the original series pulled as-is: PTAX, cumulative IPCA index, "
         "cumulative US CPI index, carry (Selic-Fed Funds), terms of trade (Funcex), "
         "fiscal risk (5Y CDS), 10y bond-implied breakeven inflation, the CMN inflation "
         "target, and the breakeven-minus-target gap. Static values, no formulas.", "body"),
        ("Why the IPCA/CPI index columns don't start at 100", "h2"),
        ("Both cumulative indices are seeded at 100 in 1994-01 (a few months before PTAX "
         "data exists, so this sheet's first visible row is 1994-07). Brazil was still in "
         "hyperinflation before the Real Plan (launched July 1994) — monthly IPCA readings "
         "in early 1994 ran ~40-50% — so by the time PTAX data starts and the row becomes "
         "visible here, six months of compounding at those rates has already pushed the "
         "index to ~916. That's real history, not a data error: the index's absolute LEVEL "
         "carries no meaning on its own, only the RATIO between two dates does "
         "(index(t)/index(base)) — which is exactly what the equilibrium formula on the "
         "next sheet uses.", "body"),
        ("PPP Equilibrium — equilibrium(t) = PTAX(base) x [IPCA(t)/IPCA(base)] / "
         "[CPI(t)/CPI(base)], and deviation(t) = 100 x LN(PTAX(t)/equilibrium(t)). "
         "The base month is a validated dropdown (cell I1) — change it and every "
         "formula recalculates. Base-month choice shifts the equilibrium level and "
         "the deviation's baseline, but not its month-to-month shape.", "body"),
        ("Model Parameters — top block: the 5 posterior means (hardcoded, labeled with "
         "source). Bottom block: standardization stats (mean/stdev of each lagged delta), "
         "computed with real AVERAGE/STDEV.S formulas over rows "
         f"{primary_start_row}-{primary_end_row} ({primary_start_date:%Y-%m} to "
         f"{primary_end_date:%Y-%m}) — the exact sample the model was fit on, "
         "bounded by fiscal-risk (CDS) data availability plus the differencing/lag.", "body"),
        ("Deltas — delta_dev = this month's deviation minus last month's; delta_x = "
         "this month's raw value minus last month's, for each of the 4 regressors; "
         "delta_x_lag1 = last month's delta_x (the model explains this month's move "
         "using LAST month's change in each regressor, since news takes time to show "
         "up); z_x = the standardized version, (delta_x_lag1 - mean)/stdev, populated "
         f"only for rows {primary_start_row}-{primary_end_row} (the model's actual sample).", "body"),
        ("Decomposition & Bridge — contribution_x(t) = beta_x * z_x(t); fitted delta_dev "
         "= alpha + sum of contributions; residual = actual - fitted. Columns M-R cumulate "
         "each piece over time (running sums); columns S-Y convert those into a running R$ "
         "level (lvl0..lvl6, the same sequential/telescoping construction as the Python "
         "code); columns Z-AE take consecutive differences of that running level to get the "
         "actual bridge contributions the dashboard's 'Nominal Exchange Rate Decomposition' "
         "chart plots — equilibrium + baseline + carry + terms of trade + breakeven + fiscal "
         "+ residual = the actual PTAX rate, exactly (column AG checks the gap, which should "
         "be ~0 up to floating-point rounding).", "body"),
        ("", ""),
        ("Why the sample starts in 2008-02", "h2"),
        ("Fiscal risk (5Y CDS, macro_brasil.cmb_risco_pais) only starts in 2007-12 — "
         "the latest-starting of the 4 regressors. Once you difference (needs t-1) and "
         "then lag by 1 more month (needs t-2), the first usable observation is 2008-02. "
         "The sample ends in 2026-05 rather than 2026-06 because terms-of-trade's most "
         "recent print (2026-04) hasn't been differenced yet as of 2026-06.", "body"),
    ]
    r = 1
    for text, kind in lines:
        cell = ws.cell(row=r, column=1, value=text)
        if kind == "title":
            cell.font = Font(bold=True, size=15, color=_NAVY)
        elif kind == "h2":
            cell.font = Font(bold=True, size=12, color=_TEAL)
        else:
            cell.font = Font(size=10.5, color="222222")
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = 18 if kind in ("title", "h2", "") else 34
        r += 1
    ws.column_dimensions["A"].width = 118

    # =====================================================================
    # Raw Data
    # =====================================================================
    ws = wb.create_sheet("Raw Data")
    headers = ["Date", "PTAX (venda)", "IPCA cumulative index (see README re: level)",
               "US CPI cumulative index (see README re: level)",
               "Carry (Selic-FedFunds, %)", "Terms of trade (Funcex idx)",
               "Fiscal risk (5Y CDS, bps)", "Breakeven infl. (10y, %)",
               "Inflation target (%, annual)", "Breakeven - target gap (%)"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=1, column=c, value=h)
    _style_header(ws, 1, len(headers))
    cols = ["ptax", "ipca_index", "cpi_index", "carry", "tot", "fiscal", "target", "breakeven_gap"]
    # column order in sheet: ptax, ipca, cpi, carry, tot, fiscal, breakeven, target, gap
    sheet_cols = ["ptax", "ipca_index", "cpi_index", "carry", "tot", "fiscal", "breakeven", "target", "breakeven_gap"]
    for i, (date, row) in enumerate(df.iterrows()):
        r = first_row + i
        ws.cell(row=r, column=1, value=date.to_pydatetime()).number_format = "yyyy-mm"
        for c, col in enumerate(sheet_cols, start=2):
            v = row[col]
            if pd.notna(v):
                ws.cell(row=r, column=c, value=round(float(v), 6))
    for c in range(2, 5):
        for r in range(first_row, last_row + 1):
            ws.cell(row=r, column=c).number_format = "0.0000"
    for c in range(5, 11):
        for r in range(first_row, last_row + 1):
            ws.cell(row=r, column=c).number_format = "0.0000"
    ws.freeze_panes = "B2"
    _autosize(ws, [11, 13, 17, 17, 15, 16, 14, 14, 15, 16])

    # =====================================================================
    # PPP Equilibrium
    # =====================================================================
    ws = wb.create_sheet("PPP Equilibrium")
    headers = ["Date", "PTAX", "IPCA index", "US CPI index", "Equilibrium (PPP-implied)", "Deviation D(t), %"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=1, column=c, value=h)
    _style_header(ws, 1, len(headers))

    ws.cell(row=1, column=8, value="Base month selector").font = Font(bold=True, color=_NAVY, size=10)
    ws.cell(row=2, column=8, value="Base month (pick any month present in Raw Data):")
    base_cell = ws.cell(row=2, column=9, value=pd.Timestamp(_DEFAULT_BASE_MONTH + "-01").to_pydatetime())
    base_cell.number_format = "yyyy-mm"
    base_cell.fill = _PARAM_FILL
    base_cell.font = Font(bold=True, color=_NAVY)
    dv = DataValidation(type="list", formula1=f"='Raw Data'!$A${first_row}:$A${last_row}", allow_blank=False)
    dv.error = "Pick a month that exists in Raw Data (col A)."
    dv.errorTitle = "Invalid base month"
    ws.add_data_validation(dv)
    dv.add(base_cell)

    ws.cell(row=3, column=8, value="Base PTAX =")
    ws.cell(row=3, column=9, value=f"=INDEX('Raw Data'!$B${first_row}:$B${last_row},MATCH($I$2,'Raw Data'!$A${first_row}:$A${last_row},0))")
    ws.cell(row=4, column=8, value="Base IPCA index =")
    ws.cell(row=4, column=9, value=f"=INDEX('Raw Data'!$C${first_row}:$C${last_row},MATCH($I$2,'Raw Data'!$A${first_row}:$A${last_row},0))")
    ws.cell(row=5, column=8, value="Base CPI index =")
    ws.cell(row=5, column=9, value=f"=INDEX('Raw Data'!$D${first_row}:$D${last_row},MATCH($I$2,'Raw Data'!$A${first_row}:$A${last_row},0))")
    for rr in (3, 4, 5):
        ws.cell(row=rr, column=9).number_format = "0.0000"

    for i in range(n):
        r = first_row + i
        ws.cell(row=r, column=1, value=f"='Raw Data'!A{r}").number_format = "yyyy-mm"
        ws.cell(row=r, column=2, value=f"='Raw Data'!B{r}")
        ws.cell(row=r, column=3, value=f"='Raw Data'!C{r}")
        ws.cell(row=r, column=4, value=f"='Raw Data'!D{r}")
        ws.cell(row=r, column=5, value=f"=IF(AND(ISNUMBER(C{r}),ISNUMBER(D{r})),$I$3*(C{r}/$I$4)/(D{r}/$I$5),\"\")")
        ws.cell(row=r, column=6, value=f"=IF(AND(ISNUMBER(B{r}),ISNUMBER(E{r}),E{r}<>\"\"),100*LN(B{r}/E{r}),\"\")")
        for c in (2, 3, 4, 5):
            ws.cell(row=r, column=c).number_format = "0.0000"
        ws.cell(row=r, column=6).number_format = "0.00"
    ws.freeze_panes = "B2"
    _autosize(ws, [11, 11, 12, 12, 20, 16, 2, 40, 12])

    # =====================================================================
    # Model Parameters
    # =====================================================================
    ws = wb.create_sheet("Model Parameters")
    ws.cell(row=1, column=1, value="Bayesian posterior means — primary_breakeven spec").font = Font(bold=True, size=12, color=_NAVY)
    ws.cell(row=2, column=1, value="(plugged in, not re-estimated here — see README)").font = Font(italic=True, size=9.5, color=_GREY)
    hdr = ["Parameter", "Posterior mean", "Source"]
    for c, h in enumerate(hdr, start=1):
        ws.cell(row=3, column=c, value=h)
    _style_header(ws, 3, len(hdr))
    src = "bayesian_results/primary_breakeven_summary.csv (saved PyMC/NUTS fit)"
    param_rows = [
        ("alpha (intercept)", _POSTERIOR["alpha"]),
        ("beta_carry", _POSTERIOR["beta_carry"]),
        ("beta_tot", _POSTERIOR["beta_tot"]),
        ("beta_breakeven", _POSTERIOR["beta_breakeven"]),
        ("beta_fiscal", _POSTERIOR["beta_fiscal"]),
    ]
    for i, (name, val) in enumerate(param_rows):
        r = 4 + i
        ws.cell(row=r, column=1, value=name)
        c = ws.cell(row=r, column=2, value=val)
        c.number_format = "0.000"
        c.fill = _PARAM_FILL
        c.font = Font(bold=True, color=_NAVY)
        ws.cell(row=r, column=3, value=src).font = Font(size=9, color=_GREY, italic=True)

    ws.cell(row=10, column=1, value=f"Standardization stats — computed over rows {primary_start_row}-{primary_end_row}").font = Font(bold=True, size=11, color=_TEAL)
    ws.cell(row=11, column=1, value=f"({primary_start_date:%Y-%m} to {primary_end_date:%Y-%m}, the exact sample the model above was fit on)").font = Font(italic=True, size=9.5, color=_GREY)
    hdr2 = ["Variable", "Mean (Deltas!)", "Stdev.S (Deltas!)"]
    for c, h in enumerate(hdr2, start=1):
        ws.cell(row=12, column=c, value=h)
    _style_header(ws, 12, len(hdr2))
    stat_specs = [
        ("delta_carry (lag-1)", "F"),
        ("delta_tot (lag-1)", "J"),
        ("delta_breakeven (lag-1)", "N"),
        ("delta_fiscal (lag-1)", "R"),
    ]
    for i, (name, col_letter) in enumerate(stat_specs):
        r = 13 + i
        ws.cell(row=r, column=1, value=name)
        ws.cell(row=r, column=2, value=f"=AVERAGE(Deltas!{col_letter}{primary_start_row}:{col_letter}{primary_end_row})").number_format = "0.0000"
        ws.cell(row=r, column=3, value=f"=STDEV.S(Deltas!{col_letter}{primary_start_row}:{col_letter}{primary_end_row})").number_format = "0.0000"
    _autosize(ws, [26, 20, 55])

    # cell refs used by Deltas + Decomposition sheets
    P = {
        "alpha": "'Model Parameters'!$B$4",
        "beta_carry": "'Model Parameters'!$B$5",
        "beta_tot": "'Model Parameters'!$B$6",
        "beta_breakeven": "'Model Parameters'!$B$7",
        "beta_fiscal": "'Model Parameters'!$B$8",
        "mean_carry": "'Model Parameters'!$B$13", "sd_carry": "'Model Parameters'!$C$13",
        "mean_tot": "'Model Parameters'!$B$14", "sd_tot": "'Model Parameters'!$C$14",
        "mean_breakeven": "'Model Parameters'!$B$15", "sd_breakeven": "'Model Parameters'!$C$15",
        "mean_fiscal": "'Model Parameters'!$B$16", "sd_fiscal": "'Model Parameters'!$C$16",
    }

    # =====================================================================
    # Deltas
    # =====================================================================
    ws = wb.create_sheet("Deltas")
    headers = ["Date", "Deviation D(t)", "delta_dev",
               "Carry", "delta_carry", "delta_carry_lag1", "z_carry",
               "ToT", "delta_tot", "delta_tot_lag1", "z_tot",
               "Breakeven", "delta_breakeven", "delta_breakeven_lag1", "z_breakeven",
               "Fiscal", "delta_fiscal", "delta_fiscal_lag1", "z_fiscal"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=1, column=c, value=h)
    _style_header(ws, 1, len(headers))

    def diff_formula(raw_col, r):
        return f"=IF(AND(ISNUMBER('Raw Data'!{raw_col}{r}),ISNUMBER('Raw Data'!{raw_col}{r - 1})),'Raw Data'!{raw_col}{r}-'Raw Data'!{raw_col}{r - 1},\"\")"

    def lag_formula(delta_col, r):
        return f"=IF(ISNUMBER({delta_col}{r - 1}),{delta_col}{r - 1},\"\")"

    def z_formula(lag_col, r, mean_ref, sd_ref):
        return f"=IF(ISNUMBER({lag_col}{r}),({lag_col}{r}-{mean_ref})/{sd_ref},\"\")"

    for i in range(n):
        r = first_row + i
        ws.cell(row=r, column=1, value=f"='Raw Data'!A{r}").number_format = "yyyy-mm"
        ws.cell(row=r, column=2, value=f"='PPP Equilibrium'!F{r}")
        if r > first_row:
            ws.cell(row=r, column=3, value=f"=IF(AND(ISNUMBER(B{r}),ISNUMBER(B{r - 1})),B{r}-B{r - 1},\"\")")

        ws.cell(row=r, column=4, value=f"='Raw Data'!E{r}")
        ws.cell(row=r, column=8, value=f"='Raw Data'!F{r}")
        ws.cell(row=r, column=12, value=f"='Raw Data'!H{r}")
        ws.cell(row=r, column=16, value=f"='Raw Data'!G{r}")
        if r > first_row:
            ws.cell(row=r, column=5, value=diff_formula("E", r))
            ws.cell(row=r, column=9, value=diff_formula("F", r))
            ws.cell(row=r, column=13, value=diff_formula("H", r))
            ws.cell(row=r, column=17, value=diff_formula("G", r))
        if r > first_row + 1:
            ws.cell(row=r, column=6, value=lag_formula("E", r))
            ws.cell(row=r, column=10, value=lag_formula("I", r))
            ws.cell(row=r, column=14, value=lag_formula("M", r))
            ws.cell(row=r, column=18, value=lag_formula("Q", r))
        if primary_start_row <= r <= primary_end_row:
            ws.cell(row=r, column=7, value=z_formula("F", r, P["mean_carry"], P["sd_carry"]))
            ws.cell(row=r, column=11, value=z_formula("J", r, P["mean_tot"], P["sd_tot"]))
            ws.cell(row=r, column=15, value=z_formula("N", r, P["mean_breakeven"], P["sd_breakeven"]))
            ws.cell(row=r, column=19, value=z_formula("R", r, P["mean_fiscal"], P["sd_fiscal"]))
        for c in (2, 3, 4, 5, 6, 8, 9, 10, 12, 13, 14, 16, 17, 18):
            ws.cell(row=r, column=c).number_format = "0.0000"
        for c in (7, 11, 15, 19):
            ws.cell(row=r, column=c).number_format = "0.000"
    ws.freeze_panes = "B2"
    _autosize(ws, [11] + [15] * 18)

    # highlight the in-sample band on column A
    for r in (primary_start_row, primary_end_row):
        ws.cell(row=r, column=1).fill = PatternFill("solid", fgColor="EAF2F6")

    # =====================================================================
    # Decomposition & Bridge
    # =====================================================================
    ws = wb.create_sheet("Decomposition & Bridge")
    ws.cell(row=1, column=1,
            value="Every row below is one month of the model's actual sample "
                  f"({primary_start_date:%Y-%m} to {primary_end_date:%Y-%m}) — this sheet "
                  "does NOT share the row-per-month backbone the other sheets use, since "
                  "it only ever covers this sample; column A still shows the real date, "
                  "and every formula pulls from the correct backbone row on the other "
                  "sheets. See README.").font = Font(italic=True, size=9.5, color=_GREY)
    ws.cell(row=1, column=11, value="Anchor: deviation one month before sample start").font = Font(bold=True, size=9.5, color=_TEAL)
    ws.cell(row=1, column=12, value=f"='PPP Equilibrium'!F{anchor_row}").number_format = "0.0000"
    anchor_ref = "$L$1"

    headers = ["Date",
               "z_carry", "z_tot", "z_breakeven", "z_fiscal",
               "contrib_carry", "contrib_tot", "contrib_breakeven", "contrib_fiscal",
               "fitted delta_dev", "actual delta_dev", "residual",
               "cum_alpha", "cum_carry", "cum_tot", "cum_breakeven", "cum_fiscal", "cum_residual",
               "equilibrium (R$) [lvl0]",
               "level after baseline [lvl1]", "level after carry [lvl2]", "level after tot [lvl3]",
               "level after breakeven [lvl4]", "level after fiscal [lvl5]", "level after residual [lvl6]",
               "bridge: baseline (R$)", "bridge: carry (R$)", "bridge: terms of trade (R$)",
               "bridge: breakeven (R$)", "bridge: fiscal (R$)", "bridge: residual (R$)",
               "actual PTAX (R$)", "check (lvl6 - actual, should be ~0)"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=2, column=c, value=h)
    _style_header(ws, 2, len(headers))

    n_sample = primary_end_row - primary_start_row + 1
    for i in range(n_sample):
        r = 3 + i                          # this sheet's own row (no gap, starts right after header)
        src = primary_start_row + i        # the matching row on Raw Data / PPP Equilibrium / Deltas
        ws.cell(row=r, column=1, value=f"='Raw Data'!A{src}").number_format = "yyyy-mm"
        ws.cell(row=r, column=2, value=f"=Deltas!G{src}")
        ws.cell(row=r, column=3, value=f"=Deltas!K{src}")
        ws.cell(row=r, column=4, value=f"=Deltas!O{src}")
        ws.cell(row=r, column=5, value=f"=Deltas!S{src}")
        ws.cell(row=r, column=6, value=f"={P['beta_carry']}*B{r}")
        ws.cell(row=r, column=7, value=f"={P['beta_tot']}*C{r}")
        ws.cell(row=r, column=8, value=f"={P['beta_breakeven']}*D{r}")
        ws.cell(row=r, column=9, value=f"={P['beta_fiscal']}*E{r}")
        ws.cell(row=r, column=10, value=f"={P['alpha']}+F{r}+G{r}+H{r}+I{r}")
        ws.cell(row=r, column=11, value=f"=Deltas!C{src}")
        ws.cell(row=r, column=12, value=f"=K{r}-J{r}")

        if i == 0:
            ws.cell(row=r, column=13, value=f"={P['alpha']}")
            ws.cell(row=r, column=14, value=f"=F{r}")
            ws.cell(row=r, column=15, value=f"=G{r}")
            ws.cell(row=r, column=16, value=f"=H{r}")
            ws.cell(row=r, column=17, value=f"=I{r}")
            ws.cell(row=r, column=18, value=f"=L{r}")
        else:
            ws.cell(row=r, column=13, value=f"=M{r - 1}+{P['alpha']}")
            ws.cell(row=r, column=14, value=f"=N{r - 1}+F{r}")
            ws.cell(row=r, column=15, value=f"=O{r - 1}+G{r}")
            ws.cell(row=r, column=16, value=f"=P{r - 1}+H{r}")
            ws.cell(row=r, column=17, value=f"=Q{r - 1}+I{r}")
            ws.cell(row=r, column=18, value=f"=R{r - 1}+L{r}")

        # cumulative levels, lvl0..lvl6 — same sequential/telescoping construction
        # as bayesian_deviation_model.py's level_decomposition block
        ws.cell(row=r, column=19, value=f"='PPP Equilibrium'!E{src}")
        ws.cell(row=r, column=20, value=f"=S{r}*EXP(({anchor_ref}+M{r})/100)")
        ws.cell(row=r, column=21, value=f"=T{r}*EXP(N{r}/100)")
        ws.cell(row=r, column=22, value=f"=U{r}*EXP(O{r}/100)")
        ws.cell(row=r, column=23, value=f"=V{r}*EXP(P{r}/100)")
        ws.cell(row=r, column=24, value=f"=W{r}*EXP(Q{r}/100)")
        ws.cell(row=r, column=25, value=f"=X{r}*EXP(R{r}/100)")

        # bridge increments (what the dashboard's "Nominal Exchange Rate
        # Decomposition" chart actually plots) = consecutive lvl differences
        ws.cell(row=r, column=26, value=f"=T{r}-S{r}")
        ws.cell(row=r, column=27, value=f"=U{r}-T{r}")
        ws.cell(row=r, column=28, value=f"=V{r}-U{r}")
        ws.cell(row=r, column=29, value=f"=W{r}-V{r}")
        ws.cell(row=r, column=30, value=f"=X{r}-W{r}")
        ws.cell(row=r, column=31, value=f"=Y{r}-X{r}")

        ws.cell(row=r, column=32, value=f"='Raw Data'!B{src}")
        ws.cell(row=r, column=33, value=f"=Y{r}-AF{r}")

        for c in list(range(2, 18)):
            ws.cell(row=r, column=c).number_format = "0.0000"
        for c in list(range(19, 34)):
            ws.cell(row=r, column=c).number_format = "0.0000"
    ws.freeze_panes = "B3"
    _autosize(ws, [11] + [15] * 32)

    return wb


def run() -> None:
    wb = build_workbook()
    _OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(_OUTPUT)
    print(f"Audit workbook written to {_OUTPUT}")


if __name__ == "__main__":
    run()
